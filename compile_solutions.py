"""
compile_solutions.py — Compile solution JSON files into an Excel summary
=========================================================================
Reads all *.json files in a solutions directory (default: "solutions/"),
excluding the "oracle_<instance>.json" hindsight-cache files (those are not
runs, they are the shared oracle result reused by every method on that
instance). Cross-references "logs/" so that runs which were started but
never produced a solution JSON — because the solver proved infeasibility, or
the process crashed/timed out/was interrupted — still show up instead of
silently disappearing.

Writes a formatted Excel workbook with:

  Sheet "Results" — one row per run (finished or not), columns:
    status, run_id, instance, method, n_scenarios, horizon_hours,
    delta, criterion, solve_mode,
    sim_arrival_h, duration_h, wall_clock_s,
    oracle_obj, gap_to_oracle_pct, oracle_mipgap_pct,
    oracle_feasible, oracle_optimal, oracle_status, note

  Sheet "Summary" — one row per (instance_family, method), averaging only
  the runs that finished successfully:
    instance_family, method, n_runs, n_failed,
    sim_arrival_h mean/min/max, oracle_obj mean/min/max,
    gap_to_oracle_pct mean, method_time_s mean

  "instance_family" strips the trailing "_<seed>" from the instance id, so
  e.g. "RshortCfewTtight_1", "RshortCfewTtight_2", ... "RshortCfewTtight_10"
  are all pooled into one "RshortCfewTtight" row per method — as are any
  repeated runs of the same instance+method (e.g. re-run on a different
  day).

Usage
-----
  python compile_solutions.py                        # reads ./solutions/, ./logs/
  python compile_solutions.py --dir path/to/sols     # custom solutions directory
  python compile_solutions.py --logs path/to/logs    # custom logs directory
  python compile_solutions.py --out results.xlsx     # custom output name
"""

from __future__ import annotations

import argparse
import json
import math
import os
import re
import sys

import openpyxl
from openpyxl.styles import (
    Alignment, Font, PatternFill, Border, Side,
    numbers as xl_numbers,
)
from openpyxl.utils import get_column_letter

# ── colour palette ──────────────────────────────────────────────────────────
_HEADER_BG   = "1F4E79"   # dark blue
_HEADER_FG   = "FFFFFF"   # white
_ALT_ROW_BG  = "D6E4F0"   # light blue
_SUMMARY_BG  = "E2EFDA"   # light green
_BORDER_COL  = "B8CCE4"

_INFEASIBLE_BG = "FFC7CE"   # red   — solver explicitly proved infeasible
_INFEASIBLE_FG = "9C0006"
_INCOMPLETE_BG = "FFEB9C"   # amber — run never finished, reason unknown
_INCOMPLETE_FG = "9C6500"
_ALLFAIL_BG    = "FFC7CE"   # summary row where every run in the group failed
_SOMEFAIL_BG   = "FFEB9C"   # summary row where some (not all) runs failed

_THIN = Side(style="thin", color=_BORDER_COL)
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# ── column specification for "Results" sheet ────────────────────────────────
# (header, json_path_as_tuple, number_format, width)
_COLS = [
    ("status",          ("status",),                    "@",        12),
    ("run_id",          ("run_id",),                    "@",        38),
    ("instance",        ("instance",),                  "@",        24),
    ("route_class",     ("route_class",),               "@",        11),
    ("customers_class", ("customers_class",),           "@",        13),
    ("window_class",    ("window_class",),               "@",        12),
    ("method",          ("method",),                    "@",        10),
    ("n_scenarios",     ("n_scenarios",),                "0",        11),
    ("horizon_h",       ("horizon_hours",),              "0.0",      10),
    ("delta",           ("delta",),                      "0.00",     8),
    ("criterion",       ("criterion",),                  "@",        10),
    ("solve_mode",      ("solve_mode",),                 "@",        11),
    ("sim_arrival_h",   ("sim_arrival_h",),              "0.000",    13),
    ("duration_h",      ("duration_h",),                 "0.000",    12),
    ("method_time_s",   ("wall_clock_s",),               "0.0",      13),
    ("oracle_obj",      ("oracle", "obj"),                "0.000",    12),
    ("gap_to_oracle_%", ("gap_to_oracle",),              "0.00%",    15),
    ("oracle_mipgap_%", ("oracle", "gap"),                "0.00%",    15),
    ("oracle_feasible", ("oracle", "feasible"),          "@",        13),
    ("oracle_optimal",  ("oracle", "optimal"),           "@",        13),
    ("oracle_status",   ("oracle", "status"),            "@",        14),
    ("note",            ("note",),                       "@",        50),
]

# run_id pattern produced by runner_dispatch.run_batch:
#   <instance>_<ALGO>_<YYYYMMDD>_<HHMMSS>_<idx>
_RUN_ID_RE = re.compile(
    r"^(?P<instance>.+)_(?P<algo>LA|RO|GREEDY|2SP)_"
    r"(?P<ts>\d{8}_\d{6})_(?P<idx>\d+)$"
)
_ALGO_TO_METHOD = {"GREEDY": "greedy", "LA": "LA", "RO": "RO", "2SP": "2SP"}

_STATUS_LINE_RE = re.compile(r"Status\s*:\s*(\w+)\s*\(([\d.]+)s\)")

# instance id -> instance family, e.g. "RshortCfewTtight_1" -> "RshortCfewTtight"
# (strips the trailing "_<seed>" so every seed of the same family is grouped
# together in the Summary sheet)
_INSTANCE_SEED_RE = re.compile(r"_\d+$")


def _instance_family(instance: str) -> str:
    return _INSTANCE_SEED_RE.sub("", instance or "")


# instance/family tag -> (route_class, customers_class, window_class), per
# the naming scheme in instance_io.instance_filename():
#   R{short|medium|long}C{few|medium|many}T{none|tight|medium|large}[_<seed>]
_INSTANCE_TAG_RE = re.compile(
    r"^R(?P<route>short|medium|long)"
    r"C(?P<cust>few|medium|many)"
    r"T(?P<window>none|tight|medium|large)"
    r"(?:_\d+)?$"
)


def _parse_instance_tags(instance: str) -> dict:
    m = _INSTANCE_TAG_RE.match(instance or "")
    if not m:
        return dict(route_class=None, customers_class=None, window_class=None)
    return dict(route_class=m.group("route"),
                customers_class=m.group("cust"),
                window_class=m.group("window"))


def _get(d: dict, path: tuple):
    """Safely walk a nested dict using a key-path tuple."""
    val = d
    for key in path:
        if not isinstance(val, dict):
            return None
        val = val.get(key)
    return val


def _safe_float(v):
    """Return float or None; treat inf/nan as None."""
    try:
        f = float(v)
        return None if (math.isnan(f) or math.isinf(f)) else f
    except (TypeError, ValueError):
        return None


def load_solutions(solutions_dir: str) -> list[dict]:
    """Load all *.json solution files from solutions_dir (skips oracle caches)."""
    if not os.path.isdir(solutions_dir):
        print(f"  ERROR: directory not found: '{solutions_dir}'", file=sys.stderr)
        sys.exit(1)

    paths = sorted(
        os.path.join(solutions_dir, f)
        for f in os.listdir(solutions_dir)
        if f.endswith(".json") and not f.startswith("oracle_")
    )
    if not paths:
        print(f"  WARNING: no run .json files found in '{solutions_dir}/'")

    rows = []
    for p in paths:
        try:
            with open(p, encoding="utf-8") as fh:
                data = json.load(fh)
            data["_file"] = os.path.basename(p)
            data["status"] = "OK"
            data["note"] = ""
            rows.append(data)
        except Exception as e:
            print(f"  SKIP {p}: {e}", file=sys.stderr)

    print(f"  Loaded {len(rows)} finished run(s) from '{solutions_dir}/'")
    return rows


def find_failed_runs(logs_dir: str, solutions_dir: str) -> list[dict]:
    """
    Find runs that have a log file but never produced a solution JSON —
    i.e. they were started but did not finish (proven infeasible, crashed,
    timed out, or interrupted).
    """
    if not os.path.isdir(logs_dir):
        print(f"  WARNING: logs directory not found: '{logs_dir}/' "
              f"(skipping unfinished-run detection)")
        return []

    finished_ids = set()
    if os.path.isdir(solutions_dir):
        finished_ids = {
            f[:-5] for f in os.listdir(solutions_dir)
            if f.endswith(".json") and not f.startswith("oracle_")
        }

    rows = []
    for f in sorted(os.listdir(logs_dir)):
        if not f.endswith(".txt"):
            continue
        run_id = f[:-4]
        if run_id in finished_ids:
            continue

        m = _RUN_ID_RE.match(run_id)
        if m:
            instance, method = m.group("instance"), _ALGO_TO_METHOD[m.group("algo")]
        else:
            instance, method = run_id, "UNKNOWN"

        log_path = os.path.join(logs_dir, f)
        try:
            text = open(log_path, encoding="utf-8", errors="replace").read()
        except Exception as e:
            text = ""
            print(f"  SKIP {log_path}: {e}", file=sys.stderr)

        sm = _STATUS_LINE_RE.search(text)
        solver_status = sm.group(1) if sm else None
        elapsed       = _safe_float(sm.group(2)) if sm else None

        if solver_status == "infeasible" or "no feasible solution found" in text.lower():
            status = "INFEASIBLE"
            note = "Solver proved infeasible"
            if elapsed is not None:
                note += f" after {elapsed:.1f}s"
        else:
            status = "INCOMPLETE"
            note = "Run log ends without a completion message"
            if solver_status is not None and elapsed is not None:
                note = (f"MILP reached '{solver_status}' after {elapsed:.1f}s "
                        f"but the run crashed/was interrupted before finishing")
            note += " (crash, timeout, or manual interruption; reason not captured in the log)"

        rows.append(dict(
            run_id=run_id, instance=instance, method=method,
            status=status, note=note,
            wall_clock_s=elapsed,
            oracle={},
        ))

    print(f"  Found {len(rows)} unfinished run(s) referenced in '{logs_dir}/' "
          f"with no matching solution file")
    return rows


def _annotate_instance_tags(rows: list[dict]):
    """Add route_class / customers_class / window_class keys parsed from the
    instance id (e.g. "RshortCfewTtight_1" -> short / few / tight)."""
    for rec in rows:
        rec.update(_parse_instance_tags(rec.get("instance")))


def _annotate_gap_to_oracle(rows: list[dict]):
    """Add a 'gap_to_oracle' key: (sim_arrival_h - oracle_obj) / oracle_obj."""
    for rec in rows:
        gap = None
        if rec.get("status") == "OK":
            sim = _safe_float(_get(rec, ("sim_arrival_h",)))
            ora = _safe_float(_get(rec, ("oracle", "obj")))
            if sim is not None and ora not in (None, 0):
                gap = (sim - ora) / ora
        rec["gap_to_oracle"] = gap


def _style_header(cell, bg=_HEADER_BG, fg=_HEADER_FG):
    cell.font      = Font(bold=True, color=fg, name="Arial", size=10)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _BORDER


def _style_data(cell, row_idx: int, num_fmt: str, status: str):
    if status == "INFEASIBLE":
        bg, fg = _INFEASIBLE_BG, _INFEASIBLE_FG
    elif status == "INCOMPLETE":
        bg, fg = _INCOMPLETE_BG, _INCOMPLETE_FG
    else:
        bg, fg = (_ALT_ROW_BG if row_idx % 2 == 0 else "FFFFFF"), "000000"
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.font      = Font(name="Arial", size=10, color=fg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = _BORDER
    cell.number_format = num_fmt


def build_results_sheet(ws, rows: list[dict]):
    ws.title = "Results"
    ws.freeze_panes = "A2"

    # header row
    for col_idx, (header, _, _, width) in enumerate(_COLS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        _style_header(cell)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 30

    # data rows
    for row_idx, rec in enumerate(rows, start=2):
        status = rec.get("status", "OK")
        for col_idx, (_, path, num_fmt, _) in enumerate(_COLS, start=1):
            raw = _get(rec, path)

            # coerce type
            if num_fmt in ("0", "0.0", "0.00", "0.000", "0.00%"):
                val = _safe_float(raw)
            elif num_fmt == "@":
                val = str(raw) if raw is not None else ""
            else:
                val = raw

            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            _style_data(cell, row_idx, num_fmt, status)

    # auto-filter
    ws.auto_filter.ref = (
        f"A1:{get_column_letter(len(_COLS))}{len(rows) + 1}"
    )


def build_summary_sheet(ws, rows: list[dict]):
    ws.title = "Summary"
    ws.freeze_panes = "C2"

    # aggregate: key = (instance_family, method) — every seed of the same
    # family (e.g. RshortCfewTtight_1, _2, _3, ...) is pooled together, plus
    # any repeated runs of the same instance+method.
    from collections import defaultdict
    groups: dict[tuple, list] = defaultdict(list)
    for rec in rows:
        key = (
            _instance_family(rec.get("instance", "?")),
            rec.get("method", "?"),
        )
        groups[key].append(rec)

    headers = [
        "instance_family", "route_class", "customers_class", "window_class",
        "method",
        "n_runs", "n_failed",
        "sim_arrival_h mean", "sim_arrival_h min", "sim_arrival_h max",
        "oracle_obj mean",    "oracle_obj min",    "oracle_obj max",
        "gap_to_oracle_% mean",
        "method_time_s mean",
    ]
    col_widths = [20, 11, 13, 12, 10, 8, 9, 18, 18, 18, 16, 16, 16, 20, 18]
    num_fmts   = ["@", "@", "@", "@", "@", "0", "0",
                  "0.000", "0.000", "0.000",
                  "0.000", "0.000", "0.000",
                  "0.00%",
                  "0.0"]

    for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        _style_header(cell, bg="375623", fg="FFFFFF")
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 30

    def _mean(lst): return sum(lst) / len(lst) if lst else None
    def _min(lst):  return min(lst) if lst else None
    def _max(lst):  return max(lst) if lst else None

    for ri, ((family, method), recs) in enumerate(
            sorted(groups.items()), start=2):
        ok_recs = [r for r in recs if r.get("status") == "OK"]
        n_runs   = len(recs)
        n_failed = n_runs - len(ok_recs)

        sims  = [v for v in (_safe_float(_get(r, ("sim_arrival_h",))) for r in ok_recs) if v is not None]
        orajs = [v for v in (_safe_float(_get(r, ("oracle", "obj"))) for r in ok_recs) if v is not None]
        gaps  = [v for v in (r.get("gap_to_oracle") for r in ok_recs) if v is not None]
        times = [v for v in (_safe_float(_get(r, ("wall_clock_s",))) for r in ok_recs) if v is not None]

        tags = _parse_instance_tags(family)
        data_row = [
            family, tags["route_class"], tags["customers_class"], tags["window_class"],
            method, n_runs, n_failed,
            _mean(sims), _min(sims), _max(sims),
            _mean(orajs), _min(orajs), _max(orajs),
            _mean(gaps),
            _mean(times),
        ]

        if n_failed == 0:
            bg = _SUMMARY_BG if ri % 2 == 0 else "FFFFFF"
        elif n_failed == n_runs:
            bg = _ALLFAIL_BG
        else:
            bg = _SOMEFAIL_BG

        for ci, (val, fmt) in enumerate(zip(data_row, num_fmts), start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill      = PatternFill("solid", start_color=bg)
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = _BORDER
            cell.number_format = fmt


def compile_to_excel(solutions_dir: str, logs_dir: str, output_path: str):
    rows = load_solutions(solutions_dir)
    rows += find_failed_runs(logs_dir, solutions_dir)
    _annotate_instance_tags(rows)
    _annotate_gap_to_oracle(rows)

    # stable, readable ordering: instance, then method, then run_id
    rows.sort(key=lambda r: (str(r.get("instance")), str(r.get("method")), str(r.get("run_id"))))

    wb = openpyxl.Workbook()
    ws_results = wb.active
    build_results_sheet(ws_results, rows)

    ws_summary = wb.create_sheet()
    build_summary_sheet(ws_summary, rows)

    wb.save(output_path)
    n_failed = sum(1 for r in rows if r.get("status") != "OK")
    print(f"  Excel saved : {output_path}")
    print(f"  Sheets      : Results ({len(rows)} rows, {n_failed} unfinished), Summary")


# ── CLI ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Compile solution JSON files (plus unfinished runs found "
                     "in logs/) into an Excel summary."
    )
    parser.add_argument(
        "--dir", default="solutions",
        help="Directory containing solution .json files (default: 'solutions')"
    )
    parser.add_argument(
        "--logs", default="logs",
        help="Directory containing run log .txt files, used to detect runs "
             "that never produced a solution file (default: 'logs')"
    )
    parser.add_argument(
        "--out", default="solution_summary.xlsx",
        help="Output Excel file path (default: 'solution_summary.xlsx')"
    )
    args = parser.parse_args()
    compile_to_excel(args.dir, args.logs, args.out)
