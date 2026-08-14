# -*- coding: utf-8 -*-
"""Compare the PD74720 trip report against ChargeAndBreak model assumptions."""
import re
import sys
import statistics as st
from datetime import datetime, timedelta

import openpyxl

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

F = r"c:/Users/celinep/Documents/GitHub/ChargeAndBreak/data/Trip_Report_PD74720_2025-09-22 00-00-00_2025-09-27 23-59-59.xlsx"

sys.path.insert(0, r"c:/Users/celinep/Documents/GitHub/ChargeAndBreak")
from src.settings import (V_NOM, ecr, SERVICE_TIME_H, Tb45, Tb15, Tb30,
                          Tr1, Tr2, Tdrv_cons, Tdrv_sh1, Tdrv_sh2,
                          XI_MIN, XI_MAX, TRAVEL_TIME_CV_TARGET,
                          xi_realised_sd, M_STOP_H, QUEUE_WAIT_MEAN_MIN,
                          VEH_MASS_KG)

wb = openpyxl.load_workbook(F, read_only=True, data_only=True)


def parse_dur(s):
    """'2 t 24 min' | '12 min' -> hours (float), None if empty."""
    if not s:
        return None
    h = re.search(r"(\d+)\s*t", s)
    m = re.search(r"(\d+)\s*min", s)
    return (int(h.group(1)) if h else 0) + (int(m.group(1)) if m else 0) / 60.0


def parse_km(s):
    if not s:
        return None
    m = re.search(r"([\d.]+)\s*km", str(s))
    return float(m.group(1)) if m else None


# ── customer postcodes from sheet3 (load/unload orders) ─────────────────────
orders = list(wb["sheet3"].iter_rows(values_only=True))[1:]
cust_pc = set()
for r in orders:
    for pc in (r[4], r[7]):
        if pc:
            cust_pc.add(re.sub(r"\s", "", str(pc)).upper())
print("customer postcodes from orders:", sorted(cust_pc))


def is_customer(addr):
    if not addr:
        return False
    a = re.sub(r"\s", "", addr.upper())
    return any(pc in a for pc in cust_pc)


# ── Stops sheet: legs + dwells ───────────────────────────────────────────────
rows = list(wb["Stops"].iter_rows(values_only=True))[1:]
FMT = "%d/%m/%Y %H:%M"

stops = []
for r in rows:
    typ, num, addr, dkm, cumkm, date, arr, dep, dwell = [
        (str(c).strip() if c is not None else "") for c in r[:9]]
    stops.append(dict(
        typ=typ, addr=addr, km=parse_km(dkm),
        arr=datetime.strptime(f"{date} {arr}", FMT) if arr else None,
        dep=datetime.strptime(f"{date} {dep}", FMT) if dep else None,
        dwell=parse_dur(dwell),
    ))

# stitch consecutive rows into legs: prev departure -> this arrival
legs = []
prev = None
for s in stops:
    if s["typ"] == "Start":
        prev = s
        continue
    if prev is not None and s["arr"] and prev["dep"] and s["km"] is not None:
        dt = (s["arr"] - prev["dep"]).total_seconds() / 3600.0
        if dt < 0:  # date rollover within a trip block (times lack the day)
            dt += 24.0
        legs.append(dict(frm=prev["addr"][:30], to=s["addr"][:30],
                         km=s["km"], h=dt))
    prev = s if s["dep"] or s["arr"] else prev
    if s["dep"] is None and s["typ"] == "Ankomst":
        prev = None

print("\n── LEGS (from Stops sheet) ──")
clean = []
for lg in legs:
    if lg["km"] and lg["km"] >= 5 and lg["h"] > 0:
        v = lg["km"] / lg["h"]
        tag = ""
        if v < 40:
            tag = "  <- embedded rest/ferry, EXCLUDED"
        else:
            clean.append(lg)
        print(f"  {lg['km']:6.0f} km in {lg['h']:5.2f} h -> {v:5.1f} km/h"
              f"   {lg['frm']} -> {lg['to']}{tag}")

speeds = [lg["km"] / lg["h"] for lg in clean]
xis = [(lg["km"] / V_NOM and lg["h"] / (lg["km"] / V_NOM)) for lg in clean]
w_speed = sum(lg["km"] for lg in clean) / sum(lg["h"] for lg in clean)

print(f"\n  clean legs: {len(clean)}, {sum(lg['km'] for lg in clean):.0f} km")
print(f"  leg speed: mean {st.mean(speeds):.1f}, "
      f"km-weighted {w_speed:.1f}, range {min(speeds):.1f}-{max(speeds):.1f} km/h"
      f"   | hypothesis V_NOM = {V_NOM:.0f} km/h")
print(f"  implied xi = t/t_nom(80): mean {st.mean(xis):.3f}, "
      f"sd {st.stdev(xis):.3f}, range {min(xis):.3f}-{max(xis):.3f}"
      f"   | model xi in [{XI_MIN:.3f}, {XI_MAX:.1f}], "
      f"E=1, realised sd {xi_realised_sd(TRAVEL_TIME_CV_TARGET):.3f}")

# ── dwells: customers vs breaks ──────────────────────────────────────────────
print("\n── DWELLS ──")
cust, brk = [], []
for s in stops:
    if s["dwell"] is None:
        continue
    (cust if is_customer(s["addr"]) else brk).append((s["addr"][:45], s["dwell"]))
print("  customer stops (address matches an order postcode):")
for a, d in cust:
    print(f"    {d*60:5.0f} min  {a}")
if cust:
    ds = [d for _, d in cust]
    print(f"    mean {st.mean(ds):.2f} h, range {min(ds):.2f}-{max(ds):.2f} h"
          f"   | hypothesis SERVICE_TIME_H = {SERVICE_TIME_H} h")
print("  non-customer stops (breaks / fuel / ferry):")
for a, d in brk:
    print(f"    {d*60:5.0f} min  {a}")
if brk:
    print(f"    | model breaks: 45 min (Tb45) or 15+30 split; "
          f"CS stop overhead M_STOP = {M_STOP_H*60:.0f} min; "
          f"queue mean {QUEUE_WAIT_MEAN_MIN:.0f} min")

# ── Trips sheet: daily driving, avg speed, fuel ─────────────────────────────
print("\n── TRIPS SHEET ──")
trows = list(wb["Trips"].iter_rows(values_only=True))[1:]
per_day = {}
fuel_tot = km_tot = 0.0
for r in trows:
    if not r[1]:
        continue
    t0 = datetime.strptime(str(r[1]).strip(), "%d/%m/%Y %H:%M:%S")
    km = parse_km(r[6]); fuel = parse_km(str(r[7]).replace("liter", "km"))
    drv = parse_dur(str(r[11]))
    avg = float(r[8]) if r[8] else None
    if km: km_tot += km
    if fuel: fuel_tot += fuel
    d = t0.date()
    per_day.setdefault(d, [0.0, 0.0, 0.0])
    per_day[d][0] += drv or 0
    per_day[d][1] += km or 0
    per_day[d][2] += fuel or 0
print("  date        drive_h    km   fuel_L  L/100km")
for d, (drv, km, fu) in sorted(per_day.items()):
    print(f"  {d}  {drv:6.2f}  {km:6.0f}  {fu:6.0f}   {100*fu/km if km else 0:5.1f}")
print(f"  | model daily driving caps: {Tdrv_sh1:.0f} h regular, "
      f"{Tdrv_sh2:.0f} h extended (max 2/wk); 4.5 h continuous (Tdrv_cons)")
print(f"  total: {km_tot:.0f} km, {fuel_tot:.0f} L  "
      f"-> {100*fuel_tot/km_tot:.1f} L/100km")

# diesel energy comparison
kwh_per_l = 9.96          # diesel LHV
eta_diesel = 0.42         # engine+driveline, motorway
e_fuel = fuel_tot * kwh_per_l / km_tot
e_wheel = e_fuel * eta_diesel
print(f"  fuel energy {e_fuel:.2f} kWh/km; at {eta_diesel:.0%} eff -> "
      f"{e_wheel:.2f} kWh/km at wheel+aux")
print(f"  | model ECR({V_NOM:.0f}) = {ecr(V_NOM):.2f} kWh/km "
      f"(battery-to-wheel, 40 t); ECR(70) = {ecr(70):.2f}")

# ── overnight rests (gap between Ankomst and next day's Start) ──────────────
print("\n── OVERNIGHT RESTS (arrival -> next start at same address) ──")
arrs = [s for s in stops if s["typ"] == "Ankomst" and s["arr"]]
starts = [s for s in stops if s["typ"] == "Start" and s["dep"]]
for a in arrs:
    nxt = [s for s in starts if s["dep"] and s["dep"] > a["arr"]
           and s["addr"] == a["addr"]]
    if nxt:
        gap = (min(s["dep"] for s in nxt) - a["arr"]).total_seconds() / 3600
        if 2 < gap < 24:
            print(f"  {gap:5.2f} h  at {a['addr'][:45]}")
print(f"  | model rests: Tr1 = {Tr1:.0f} h regular, Tr2 = {Tr2:.0f} h reduced "
      f"(max 3/wk)")

# ── weights from Sheet2 ─────────────────────────────────────────────────────
print("\n── WEIGHTS (Sheet2 samples) ──")
ws = wb["Sheet2"]
gross = []
for r in ws.iter_rows(values_only=True):
    if r[10]:
        try:
            gross.append(float(str(r[10]).replace(",", ".").rstrip("t")))
        except ValueError:
            pass
if gross:
    print(f"  gross weight: {len(gross)} samples, "
          f"min {min(gross):.1f} t, max {max(gross):.1f} t, "
          f"mean {st.mean(gross):.1f} t   | model assumes {VEH_MASS_KG/1000:.0f} t")
