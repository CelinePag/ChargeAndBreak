# -*- coding: utf-8 -*-
"""Reconstruct the PD74720 route purely from the trip-report xlsx.

No external APIs: GPS spine from Sheet2, km axis from the Stops sheet
(odometer-based leg distances), customer classification from sheet3 orders.

Route conventions (agreed design):
  - the 2-min depot pass-through after the local loop is dropped;
  - each ferry crossing (ports + sailing) collapses into ONE ferry node —
    in the instance it becomes a layby with a forced break of the crossing
    duration and M_maneuver = observed boarding overhead.

Importable: build_route() returns the stop structure; make_figure() renders
the overview (optionally with charger / rest-area overlays from OSM).
"""
import re
import sys
from datetime import datetime
from math import radians, sin, cos, asin, sqrt

import numpy as np
import openpyxl

sys.path.insert(0, r"c:/Users/celinep/Documents/GitHub/ChargeAndBreak")

F = (r"c:/Users/celinep/Documents/GitHub/ChargeAndBreak/data/"
     r"Trip_Report_PD74720_2025-09-22 00-00-00_2025-09-27 23-59-59.xlsx")
OUT_PNG = r"c:/Users/celinep/Documents/GitHub/ChargeAndBreak/figures/real_route_overview.png"
OUT_CSV = r"c:/Users/celinep/Documents/GitHub/ChargeAndBreak/data_output/real_route_stops.csv"

DEPOT_PAT = "4848 Arendal"
# exact terminal addresses only (the Go'on cafeteria on Dalsagervej is a
# regular break stop 11 km before the port, not the terminal itself)
FERRY_PORT_PATS = ("Havnegata 61", "Sigurd Espersens", "Dalsagervej 5")

TYPE_STYLE = {  # color (Okabe-Ito subset) + marker shape as secondary encoding
    "depot":    ("#000000", "s", "Depot (Arendal)"),
    "customer": ("#D55E00", "o", "Customer (load/unload)"),
    "break":    ("#0072B2", "v", "Break / fuel stop"),
    "rest":     ("#009E73", "^", "Overnight rest"),
    "ferry":    ("#56B4E9", "D", "Ferry node (forced break)"),
}
CS_COLOR  = "#E69F00"   # charger overlay ticks
LAY_COLOR = "#8a8a8a"   # rest-area overlay ticks


def parse_dur(s):
    if not s:
        return 0.0
    h = re.search(r"(\d+)\s*t", s)
    m = re.search(r"(\d+)\s*min", s)
    return (int(h.group(1)) if h else 0) + (int(m.group(1)) if m else 0) / 60.0


def parse_km(s):
    m = re.search(r"([\d.]+)\s*km", str(s or ""))
    return float(m.group(1)) if m else None


def haversine(lat1, lon1, lat2, lon2):
    dlat, dlon = radians(lat2 - lat1), radians(lon2 - lon1)
    a = sin(dlat / 2) ** 2 + cos(radians(lat1)) * cos(radians(lat2)) * sin(dlon / 2) ** 2
    return 2 * 6371.0 * asin(sqrt(a))


def addr_norm(a):
    return re.sub(r"\s", "", (a or "").upper())


def build_route():
    """Parse the xlsx into (visits, gps, total_km, anchors).

    visits  : list of stop dicts (km, type, arr, dep, dwell_h, lat/lon, addr)
    gps     : list of (datetime, lat, lon), 1-minute spine
    anchors : list of (gps_index, km) — every raw stop with a known arrival,
              BEFORE any merging/dropping (used to build the km-spine)
    """
    wb = openpyxl.load_workbook(F, read_only=True, data_only=True)

    # ── GPS spine ───────────────────────────────────────────────────────────
    gps = []
    for r in wb["Sheet2"].iter_rows(values_only=True):
        t, lat, lon = r[0], r[5], r[6]
        if lat in (None, "") or lon in (None, ""):
            continue
        if isinstance(t, str):
            try:
                t = datetime.strptime(t.strip(), "%Y-%m-%d %H:%M:%S")
            except ValueError:
                continue
        if not isinstance(t, datetime):
            continue
        try:
            gps.append((t, float(str(lat).replace(",", ".")),
                        float(str(lon).replace(",", "."))))
        except ValueError:
            continue
    gps.sort(key=lambda x: x[0])
    gps_t = [g[0] for g in gps]

    def gps_idx(t):
        i = min(range(len(gps)), key=lambda j: abs((gps_t[j] - t).total_seconds()))
        return i

    def pos_at(t, tol_min=20):
        i = gps_idx(t)
        if abs((gps_t[i] - t).total_seconds()) <= tol_min * 60:
            return gps[i][1], gps[i][2]
        return None, None

    # ── customer postcodes from sheet3 ──────────────────────────────────────
    cust_pc = set()
    for r in list(wb["sheet3"].iter_rows(values_only=True))[1:]:
        for pc in (r[4], r[7]):
            if pc:
                cust_pc.add(re.sub(r"\s", "", str(pc)).upper())

    def is_customer(addr):
        a = addr_norm(addr)
        return any(pc in a or (re.match(r"\d{4,}", pc)
                               and re.match(r"\d{4,}", pc).group(0) in a)
                   for pc in cust_pc)

    # ── Stops sheet -> visit timeline with global cumulative km ─────────────
    rows = list(wb["Stops"].iter_rows(values_only=True))[1:]
    FMT = "%d/%m/%Y %H:%M"
    visits, km_offset, block_last_cum = [], 0.0, 0.0
    for r in rows:
        typ, num, addr, dkm, cumkm, date, arr, dep, dwell = [
            (str(c).strip() if c is not None else "") for c in r[:9]]
        t_arr = datetime.strptime(f"{date} {arr}", FMT) if arr else None
        t_dep = datetime.strptime(f"{date} {dep}", FMT) if dep else None
        if typ == "Start":
            km_offset += block_last_cum
            block_last_cum = 0.0
            if visits and addr_norm(visits[-1]["addr"]) == addr_norm(addr):
                visits[-1]["dep"] = t_dep or visits[-1]["dep"]
                continue
            visits.append(dict(addr=addr, arr=None, dep=t_dep, km=km_offset))
            continue
        cum = parse_km(cumkm) or 0.0
        block_last_cum = max(block_last_cum, cum)
        g_km = km_offset + cum
        if visits and addr and addr_norm(visits[-1]["addr"]) == addr_norm(addr) \
                and (g_km - visits[-1]["km"]) <= 1.5:
            visits[-1]["dep"] = t_dep or visits[-1]["dep"]
            visits[-1]["km"] = g_km
            continue
        visits.append(dict(addr=addr, arr=t_arr, dep=t_dep, km=g_km))
    total_km = km_offset + block_last_cum

    # anchors for the km-spine: every raw visit with a known arrival
    anchors = [(gps_idx(v["arr"]), v["km"]) for v in visits if v["arr"]]

    # ── classify ────────────────────────────────────────────────────────────
    for v in visits:
        a = v["addr"]
        dwell_h = ((v["dep"] - v["arr"]).total_seconds() / 3600
                   if v["arr"] and v["dep"] else None)
        if not a or any(p in a for p in FERRY_PORT_PATS):
            v["type"] = "ferry"
        elif DEPOT_PAT in a:
            v["type"] = "depot"
        elif is_customer(a):
            v["type"] = "customer"
        elif dwell_h is not None and dwell_h >= 6.0:
            v["type"] = "rest"
        else:
            v["type"] = "break"
        v["dwell_h"] = dwell_h
        v["lat"], v["lon"] = (pos_at(v["arr"] or v["dep"])
                              if (v["arr"] or v["dep"]) else (None, None))

    # rests hidden inside a leg (dep -> next arr gap >> nominal drive)
    for a, b in zip(visits, visits[1:]):
        if a["dep"] and b["arr"]:
            gap = (b["arr"] - a["dep"]).total_seconds() / 3600
            if gap - (b["km"] - a["km"]) / 80.0 >= 6.0:
                a["embedded_rest_after_h"] = gap - (b["km"] - a["km"]) / 80.0

    # merge the double Bayreuth gate stop (two customer rows, <=3 km apart)
    merged = []
    for v in visits:
        if (merged and v["type"] == "customer"
                and merged[-1]["type"] == "customer"
                and v["km"] - merged[-1]["km"] <= 3.0):
            m = merged[-1]
            m["dep"] = v["dep"] or m["dep"]
            m["dwell_h"] = (m["dwell_h"] or 0) + (v["dwell_h"] or 0)
            m["addr"] += "  (+gate merge)"
            continue
        merged.append(v)
    visits = merged

    # collapse each port cluster into ONE ferry node (agreed design)
    merged = []
    for v in visits:
        if (merged and v["type"] == "ferry" and merged[-1]["type"] == "ferry"
                and v["km"] - merged[-1]["km"] <= 5.0):
            m = merged[-1]
            m["dep"] = v["dep"] or m["dep"]
            m["km"] = v["km"]
            continue
        merged.append(v)
    visits = merged

    # drop pass-through depot visits (neither endpoint, dwell < 15 min)
    visits = [v for i, v in enumerate(visits)
              if not (v["type"] == "depot" and 0 < i < len(visits) - 1
                      and (v["dwell_h"] or 0) < 0.25)]

    # Ferry node dwell = port-to-port, measured (not estimated): arrival at
    # the embarkation port -> first engine-on after disembarkation, which the
    # Trips sheet records as the start of the next trip.  (The telematics
    # stamps that trip's START POSITION at the last fix before the sea gap,
    # i.e. the departure port, while its start TIME is already on the far
    # side.)
    trip_starts = []
    for r in list(wb["Trips"].iter_rows(values_only=True))[1:]:
        if r[1]:
            try:
                trip_starts.append(datetime.strptime(
                    str(r[1]).strip(), "%d/%m/%Y %H:%M:%S"))
            except ValueError:
                pass
    trip_starts.sort()
    for v in visits:
        if v["type"] == "ferry" and v["arr"]:
            nxt = [t for t in trip_starts if t > v["arr"]]
            if nxt:
                v["dep"] = nxt[0]
                v["ferry_dwell_h"] = (nxt[0] - v["arr"]).total_seconds() / 3600

    return dict(visits=visits, gps=gps, total_km=total_km, anchors=anchors)


def spine_km(route):
    """Cumulative-km position for every GPS point, anchored piecewise to the
    odometer-based stop positions (sea segments collapse to constant km)."""
    gps, anchors = route["gps"], sorted(set(route["anchors"]))
    # keep strictly increasing (idx, km)
    seq, last_i, last_k = [], -1, -1.0
    for i, k in anchors:
        if i > last_i and k >= last_k:
            seq.append((i, k))
            last_i, last_k = i, k
    km = np.zeros(len(gps))
    lat = np.array([g[1] for g in gps])
    lon = np.array([g[2] for g in gps])
    step = np.zeros(len(gps))
    for j in range(1, len(gps)):
        step[j] = haversine(lat[j-1], lon[j-1], lat[j], lon[j])
    cum = np.cumsum(step)
    for (i0, k0), (i1, k1) in zip(seq, seq[1:]):
        span = cum[i1] - cum[i0]
        if span <= 0:
            km[i0:i1+1] = k0
        else:
            km[i0:i1+1] = k0 + (cum[i0:i1+1] - cum[i0]) / span * (k1 - k0)
    km[:seq[0][0]] = seq[0][1]
    km[seq[-1][0]:] = seq[-1][1]
    return lat, lon, km


def make_figure(route, chargers=None, restareas=None, alt_chargers=None,
                out_png=OUT_PNG):
    """Overview figure; optional overlays are lists of dicts with
    keys km, lat, lon (chargers also: kw).  alt_chargers = candidate
    >=350 kW sites not in the HGV-tagged K set."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from src.plot.paper_style import apply_rc, INK_MUTED, GRID

    visits, gps, total_km = route["visits"], route["gps"], route["total_km"]
    apply_rc()
    fig = plt.figure(figsize=(9.5, 9.8))
    gs = fig.add_gridspec(3, 1, height_ratios=[3.2, 2.2, 1.6], hspace=0.34)

    # panel A — geographic trace
    ax = fig.add_subplot(gs[0])
    ax.plot([g[2] for g in gps], [g[1] for g in gps],
            color="#b0b0b0", lw=0.9, zorder=1)
    if restareas:
        ax.scatter([p["lon"] for p in restareas], [p["lat"] for p in restareas],
                   s=4, c=LAY_COLOR, marker=".", zorder=2, label="Rest area (OSM)")
    if alt_chargers:
        ax.scatter([p["lon"] for p in alt_chargers],
                   [p["lat"] for p in alt_chargers],
                   s=18, facecolors="none", edgecolors=CS_COLOR,
                   linewidths=0.9, marker="o", zorder=2,
                   label="Candidate ≥350 kW (untagged)")
    if chargers:
        ax.scatter([p["lon"] for p in chargers], [p["lat"] for p in chargers],
                   s=14, c=CS_COLOR, marker="P", zorder=2,
                   label="HDV charger K (OSM hgv + curated)")
    for typ, (col, mk, lbl) in TYPE_STYLE.items():
        xs = [v["lon"] for v in visits if v["type"] == typ and v["lon"]]
        ys = [v["lat"] for v in visits if v["type"] == typ and v["lat"]]
        ax.scatter(xs, ys, s=42 if typ != "break" else 26, c=col, marker=mk,
                   edgecolors="white", linewidths=0.8, zorder=3, label=lbl)
    labeled = []
    for v in visits:
        if v["type"] == "customer" and v["lon"]:
            if any(abs(v["lon"] - lo) + abs(v["lat"] - la) < 0.8
                   for lo, la in labeled):
                continue
            labeled.append((v["lon"], v["lat"]))
            city = v["addr"].split(",")[-2].strip() if "," in v["addr"] else v["addr"]
            ax.annotate(re.sub(r"^[\dA-Z ]*? ", "", city), (v["lon"], v["lat"]),
                        textcoords="offset points", xytext=(7, -2), fontsize=7)
    ax.annotate("ferry\nKristiansand–Hirtshals", (7.4, 57.5), fontsize=7,
                color="#56B4E9", ha="right", style="italic")
    ax.set_aspect(1 / np.cos(np.radians(55)))
    ax.set_xlabel("Longitude"); ax.set_ylabel("Latitude")
    ax.set_title("A — GPS trace, week 22–27 Sep 2025"
                 + (" + OSM overlays" if chargers or restareas else " (no map data)"))
    ax.legend(loc="center left", bbox_to_anchor=(1.02, 0.5), frameon=False,
              fontsize=7)
    ax.grid(color=GRID, lw=0.4)

    # panel B — space-time
    ax = fig.add_subplot(gs[1])
    t0 = min(v["arr"] or v["dep"] for v in visits if v["arr"] or v["dep"])
    pts = [((t - t0).total_seconds() / 86400, v["km"])
           for v in visits for t in (v["arr"], v["dep"]) if t]
    ax.plot([p[0] for p in pts], [p[1] for p in pts],
            color="#b0b0b0", lw=1.0, zorder=1)
    for typ, (col, mk, lbl) in TYPE_STYLE.items():
        sel = [v for v in visits if v["type"] == typ and (v["arr"] or v["dep"])]
        ax.scatter([((v["arr"] or v["dep"]) - t0).total_seconds() / 86400
                    for v in sel], [v["km"] for v in sel],
                   s=34 if typ != "break" else 20, c=col, marker=mk,
                   edgecolors="white", linewidths=0.7, zorder=3)
    ax.set_xlabel("Days since departure (Mon 22/09 07:47)")
    ax.set_ylabel("Cumulative km")
    ax.set_title("B — space–time trajectory (horizontal = stopped; ferry adds 0 km)")
    ax.grid(color=GRID, lw=0.4)

    # panel C — linear instance strip
    ax = fig.add_subplot(gs[2])
    ax.axhline(0, color=INK_MUTED, lw=1.0, zorder=1)
    if restareas:
        ax.vlines([p["km"] for p in restareas], -0.55, -0.25, colors=LAY_COLOR,
                  lw=0.7, zorder=2)
    if alt_chargers:
        from src.plot.paper_style import tint
        ax.vlines([p["km"] for p in alt_chargers], 0.25, 0.42,
                  colors=[tint(CS_COLOR, 0.55)], lw=0.9, zorder=2)
    if chargers:
        ax.vlines([p["km"] for p in chargers], 0.25, 0.55, colors=CS_COLOR,
                  lw=0.9, zorder=2)
    for typ, (col, mk, lbl) in TYPE_STYLE.items():
        xs = [v["km"] for v in visits if v["type"] == typ]
        ax.scatter(xs, [0] * len(xs), s=48 if typ != "break" else 30, c=col,
                   marker=mk, edgecolors="white", linewidths=0.8, zorder=3)
    slots = [18, -34, 42]
    last_x = {s: -1e9 for s in slots}
    si = 0
    for v in visits:
        lab = None
        if v["type"] == "customer":
            city = v["addr"].split(",")[-2].strip() if "," in v["addr"] else v["addr"]
            city = re.sub(r"^[\dA-Z]{4,8}\s", "", city)
            lab = (f"{city}\n{v['dwell_h']*60:.0f} min"
                   if v["dwell_h"] else city)
        elif v["type"] == "ferry" and v.get("ferry_dwell_h"):
            lab = f"Ferry\n~{v['ferry_dwell_h']:.1f} h"
        if lab is None:
            continue
        for _ in range(len(slots)):
            if v["km"] - last_x[slots[si]] > 220:
                break
            si = (si + 1) % len(slots)
        y = slots[si]
        last_x[y] = v["km"]
        si = (si + 1) % len(slots)
        ax.annotate(lab, (v["km"], 0), textcoords="offset points",
                    xytext=(0, y), ha="center", fontsize=6.5)
    n_cs = len(chargers) if chargers else 0
    n_ra = len(restareas) if restareas else 0
    n_alt = len(alt_chargers) if alt_chargers else 0
    extra = (f"; K = {n_cs} HDV chargers, {n_alt} candidate ≥350 kW, "
             f"{n_ra} rest areas (0.5 km) — OSM"
             if (chargers or restareas) else
             " (chargers/laybys to be overlaid from OSM)")
    ax.set_ylim(-1.6, 1.6)
    ax.set_yticks([])
    ax.set_xlim(-60, total_km + 60)
    ax.set_xlabel("Route position (km)")
    ax.set_title(f"C — linear instance: {total_km:.0f} km, "
                 f"{sum(1 for v in visits if v['type'] == 'customer')} customers"
                 f"{extra}", pad=6)
    for s in ("left", "right", "top"):
        ax.spines[s].set_visible(False)

    import os
    os.makedirs(os.path.dirname(out_png), exist_ok=True)
    fig.savefig(out_png, dpi=170)
    plt.close(fig)
    return out_png


INSTANCE_PNG = (r"c:/Users/celinep/Documents/GitHub/ChargeAndBreak/"
                r"figures/real_route_instance.png")


def make_instance_figure(route, kset, laybys, out_png=INSTANCE_PNG,
                         anonymize=True):
    """The instance as fed to the model: depot, customers, ferry nodes,
    charger set K and (aggregated) layby set L only — no observed driver
    stops, no candidate layers.

    anonymize=True (paper default) replaces customer place names with
    C1..Cn and keeps the title at country/region level, so the figure
    carries no carrier or customer identity.
    """
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from src.plot.paper_style import apply_rc, INK_MUTED, GRID

    visits = [v for v in route["visits"]
              if v["type"] in ("depot", "customer", "ferry")]
    gps, total_km = route["gps"], route["total_km"]
    n_stops = len(visits) + len(kset) + len(laybys)

    apply_rc()
    fig = plt.figure(figsize=(9.5, 6.4))
    gs = fig.add_gridspec(2, 1, height_ratios=[3.2, 1.5], hspace=0.30)

    # panel A — geography
    ax = fig.add_subplot(gs[0])
    ax.plot([g[2] for g in gps], [g[1] for g in gps],
            color="#b0b0b0", lw=0.9, zorder=1, label="Route (3339 km)")
    ax.scatter([p["lon"] for p in laybys], [p["lat"] for p in laybys],
               s=8, c=LAY_COLOR, marker=".", zorder=2,
               label=f"Layby L ({len(laybys)})")
    ax.scatter([p["lon"] for p in kset], [p["lat"] for p in kset],
               s=20, c=CS_COLOR, marker="P", zorder=3,
               label=f"Charging station K ({len(kset)})")
    for typ, lbl in (("depot", "Depot (start/end)"),
                     ("customer", f"Customer C (7)"),
                     ("ferry", "Ferry node (forced break)")):
        col, mk, _ = TYPE_STYLE[typ]
        sel = [v for v in visits if v["type"] == typ and v["lon"]]
        ax.scatter([v["lon"] for v in sel], [v["lat"] for v in sel],
                   s=46, c=col, marker=mk, edgecolors="white",
                   linewidths=0.8, zorder=4, label=lbl)
    cust_seq = [v for v in visits if v["type"] == "customer"]
    cust_id = {id(v): f"C{i + 1}" for i, v in enumerate(cust_seq)}
    labeled = []
    for v in cust_seq:
        if not v["lon"]:
            continue
        near = [lb for lb in labeled
                if abs(v["lon"] - lb[0]) + abs(v["lat"] - lb[1]) < 0.8]
        if near:
            near[0][2].append(cust_id[id(v)])
            continue
        labeled.append((v["lon"], v["lat"], [cust_id[id(v)]]))
    for lo, la, ids in labeled:
        if anonymize:
            txt = ", ".join(ids)
        else:
            v0 = next(v for v in cust_seq
                      if v["lon"] == lo and v["lat"] == la)
            txt = (v0["addr"].split(",")[-2].strip()
                   if "," in v0["addr"] else v0["addr"])
            txt = re.sub(r"^[\dA-Z ]*? ", "", txt)
        ax.annotate(txt, (lo, la), textcoords="offset points",
                    xytext=(10, -2), fontsize=7)
    ax.set_aspect(1 / np.cos(np.radians(55)))
    ax.margins(x=0.16, y=0.04)          # room for the right-edge labels
    ax.set_xlabel("Longitude"); ax.set_ylabel("Latitude")
    ax.set_title("Real-life instance — southern Norway to central "
                 "Europe, one sea crossing" if anonymize else
                 "Real-life instance — Arendal round tour via "
                 "Kristiansand–Hirtshals ferry")
    ax.legend(loc="center left", bbox_to_anchor=(1.02, 0.5), frameon=False,
              fontsize=7)
    ax.grid(color=GRID, lw=0.4)

    # panel B — the linear instance
    ax = fig.add_subplot(gs[1])
    ax.axhline(0, color=INK_MUTED, lw=1.0, zorder=1)
    ax.vlines([p["km"] for p in laybys], -0.52, -0.22, colors=LAY_COLOR,
              lw=0.8, zorder=2)
    ax.vlines([p["km"] for p in kset], 0.22, 0.52, colors=CS_COLOR,
              lw=1.0, zorder=2)
    for typ in ("depot", "customer", "ferry"):
        col, mk, _ = TYPE_STYLE[typ]
        xs = [v["km"] for v in visits if v["type"] == typ]
        ax.scatter(xs, [0] * len(xs), s=52, c=col, marker=mk,
                   edgecolors="white", linewidths=0.8, zorder=3)
    slots = [16, -34, 42]
    last_x = {s: -1e9 for s in slots}
    si = 0
    for v in visits:
        lab = None
        if v["type"] == "customer":
            if anonymize:
                city = cust_id[id(v)]
            else:
                city = (v["addr"].split(",")[-2].strip()
                        if "," in v["addr"] else v["addr"])
                city = re.sub(r"^[\dA-Z]{4,8}\s", "", city)
            lab = (f"{city}\n{v['dwell_h']*60:.0f} min"
                   if v["dwell_h"] else city)
        elif v["type"] == "ferry" and v.get("ferry_dwell_h"):
            lab = f"Ferry\n{v['ferry_dwell_h']:.1f} h"
        if lab is None:
            continue
        for _ in range(len(slots)):
            if v["km"] - last_x[slots[si]] > 220:
                break
            si = (si + 1) % len(slots)
        y = slots[si]
        last_x[y] = v["km"]
        si = (si + 1) % len(slots)
        ax.annotate(lab, (v["km"], 0), textcoords="offset points",
                    xytext=(0, y), ha="center", fontsize=6.5)
    ax.set_ylim(-1.6, 1.6)
    ax.set_yticks([])
    ax.set_xlim(-60, total_km + 60)
    ax.set_xlabel("Route position (km)")
    ax.set_title(f"Linear instance fed to the model — {n_stops} stops: "
                 f"7 customers, {len(kset)} CS, {len(laybys)} laybys, "
                 f"2 ferry nodes, depot at both ends", pad=6)
    for s in ("left", "right", "top"):
        ax.spines[s].set_visible(False)

    import os
    os.makedirs(os.path.dirname(out_png), exist_ok=True)
    fig.savefig(out_png, dpi=170)
    plt.close(fig)
    return out_png


def print_table(route):
    visits = route["visits"]
    print(f"\n{'#':>2} {'km':>6}  {'type':8} {'arrive':11}  {'dwell':>6}  address")
    for i, v in enumerate(visits):
        arr = f"{v['arr']:%d/%m %H:%M}" if v["arr"] else "-"
        if v["type"] == "ferry" and v.get("ferry_dwell_h"):
            dw = f"~{v['ferry_dwell_h']:.1f}h"
        elif v["dwell_h"] is not None:
            dw = f"{v['dwell_h']*60:4.0f}m"
        else:
            dw = "-"
        er = v.get("embedded_rest_after_h")
        note = f"  [+{er:.1f} h rest on next leg]" if er else ""
        print(f"{i:2d} {v['km']:6.0f}  {v['type']:8} {arr:11}  {dw:>6}  "
              f"{v['addr'][:52]}{note}")


def write_csv(route, out_csv=OUT_CSV):
    import os
    lines = ["idx,km,type,arrive,depart,dwell_h,ferry_dwell_h,"
             "embedded_rest_h,lat,lon,address"]
    for i, v in enumerate(route["visits"]):
        lines.append(",".join([
            str(i), f"{v['km']:.1f}", v["type"],
            f"{v['arr']:%Y-%m-%d %H:%M}" if v["arr"] else "",
            f"{v['dep']:%Y-%m-%d %H:%M}" if v["dep"] else "",
            f"{v['dwell_h']:.3f}" if v["dwell_h"] is not None else "",
            f"{v['ferry_dwell_h']:.2f}" if v.get("ferry_dwell_h") else "",
            f"{v['embedded_rest_after_h']:.2f}"
            if v.get("embedded_rest_after_h") else "",
            f"{v['lat']:.5f}" if v["lat"] else "",
            f"{v['lon']:.5f}" if v["lon"] else "",
            '"' + v["addr"].replace('"', "'") + '"']))
    os.makedirs(os.path.dirname(out_csv), exist_ok=True)
    with open(out_csv, "w", encoding="utf-8") as fh:
        fh.write("\n".join(lines))


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    route = build_route()
    print(f"GPS spine: {len(route['gps'])} points; "
          f"{len(route['visits'])} stops, total {route['total_km']:.0f} km")
    print_table(route)
    write_csv(route)
    print(f"\nfigure -> {make_figure(route)}")
    print(f"csv    -> {OUT_CSV}")
