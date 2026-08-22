"""
compile_solutions.py — Compile solution JSON files into an Excel summary
=========================================================================
Reads all *.json files in a solutions directory (default: "solutions/"),
excluding the "oracle_<instance>.json" hindsight-cache files (those are not
runs, they are the shared oracle result reused by every method on that
instance). Cross-references "logs/" so that runs which were started but
never produced a solution JSON — because the solver proved infeasibility, or
the process crashed/timed out/was interrupted — still show up instead of
silently disappearing.

Writes a formatted Excel workbook with:

  Sheet "Results" — one row per run (finished or not); every scalar field now
    stored in the solution JSON, including the full ``metrics`` block
    (HoS violations, stranding, repairs, supervisor interventions, time-window
    hit rate + penalty, per-stop decision times, offline solve time, directive
    compliance) and the per-method solve objects (RO gamma/obj, 2SP obj, ...).

  Sheet "Summary" — one row per (instance_family, method), pooling every seed
    of the same family (route × customers × window class), with means/rates for
    every metric.

  Sheet "LaTeX_Gap"        — Table 1: UNSUPERVISED gap to oracle (%), broken
    down by Route × Customers × Time-Window class × method (ready to paste).
  Sheet "LaTeX_Gap_Sup"    — same table for SUPERVISED runs (only emitted when
    supervised runs exist).
  Sheet "LaTeX_Feasibility" — Table 2: feasibility / robustness by method.
  Sheet "LaTeX_Runtime"     — Table 3: offline solve + per-stop decision times
    by method and instance size class.

Each "LaTeX_*" sheet holds the raw LaTeX source, one line per cell down
column A — select the column, copy, and paste straight into the paper.

  "instance_family" strips the trailing "_<seed>" from the instance id, so
  e.g. "RshortCfewTtight_1", "RshortCfewTtight_2", ... are all pooled into one
  "RshortCfewTtight" row per method.

Usage
-----
  python -m src.output_analysis.compile_solutions                        # reads ./solutions/, ./logs/
  python -m src.output_analysis.compile_solutions --dir path/to/sols     # custom solutions directory
  python -m src.output_analysis.compile_solutions --logs path/to/logs    # custom logs directory
  python -m src.output_analysis.compile_solutions --out results.xlsx     # custom output name
  python -m src.output_analysis.compile_solutions --tex-dir ''          # skip the .tex dump
"""

from __future__ import annotations

import argparse
import json
import math
import os
import re
import sys
from collections import Counter, defaultdict

import openpyxl
from openpyxl.styles import (
    Alignment, Font, PatternFill, Border, Side,
)
from openpyxl.utils import get_column_letter
from src import paths as _paths
from src.output_analysis import run_cache

# ── colour palette ──────────────────────────────────────────────────────────
_HEADER_BG   = "1F4E79"   # dark blue
_HEADER_FG   = "FFFFFF"   # white
_ALT_ROW_BG  = "D6E4F0"   # light blue
_SUMMARY_BG  = "E2EFDA"   # light green
_BORDER_COL  = "B8CCE4"

_INFEASIBLE_BG = "FFC7CE"   # red   — true infeasibility (certified plan fails)
_INFEASIBLE_FG = "9C0006"
_INCOMPLETE_BG = "FFEB9C"   # amber — run never finished, reason unknown
_INCOMPLETE_FG = "9C6500"
_UNSOLVED_BG   = "FCE4D6"   # peach — plan not certified (solver did not finish)
_UNSOLVED_FG   = "833C00"
_ALLFAIL_BG    = "FFC7CE"   # summary row where every run in the group failed
_SOMEFAIL_BG   = "FFEB9C"   # summary row where some (not all) runs failed

_THIN = Side(style="thin", color=_BORDER_COL)
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# ── column specification for "Results" sheet ────────────────────────────────
# (header, json_path_as_tuple, number_format, width).  num_fmt "@" => text,
# anything else => numeric (percent formats show the stored fraction as %).
_COLS = [
    ("status",           ("status",),                                     "@",     12),
    ("outcome",          ("outcome",),                                    "@",     14),
    ("solve_status",     ("solve_status",),                               "@",     12),
    ("run_id",           ("run_id",),                                     "@",     38),
    ("instance",         ("instance",),                                   "@",     22),
    ("route_class",      ("route_class",),                                "@",     11),
    ("customers_class",  ("customers_class",),                            "@",     13),
    ("window_class",     ("window_class",),                               "@",     12),
    ("method",           ("method",),                                     "@",     9),
    ("variant",          ("variant",),                                    "@",     10),
    ("criterion",        ("criterion",),                                  "@",     9),
    ("n_scenarios",      ("n_scenarios",),                                "0",     11),
    ("horizon_h",        ("horizon_hours",),                              "0.0",   10),
    ("cv",               ("cv",),                                         "0.00",  7),
    ("gamma",            ("gamma",),                                      "0.0",   7),
    ("solve_mode",       ("solve_mode",),                                 "@",     10),
    ("prune_quantile",   ("prune_quantile",),                             "0.00",  9),
    ("sim_arrival_h",    ("sim_arrival_h",),                              "0.000", 13),
    ("duration_h",       ("duration_h",),                                 "0.000", 12),
    ("duration_pen_h",   ("duration_pen_h",),                             "0.000", 13),
    ("method_time_s",    ("wall_clock_s",),                               "0.0",   13),
    ("offline_solve_s",  ("metrics", "offline_solve_time_s"),            "0.0",   14),
    ("decision_mean_s",  ("metrics", "decision_time_mean_s"),            "0.0000", 14),
    ("decision_max_s",   ("metrics", "decision_time_max_s"),             "0.0000", 13),
    ("oracle_dur_h",     ("oracle_duration_h",),                          "0.000", 12),
    ("oracle_dur_pen_h", ("oracle_duration_pen_h",),                      "0.000", 14),
    ("oracle_tw_misses", ("oracle_tw_misses",),                           "0",     13),
    ("gap_pen_%",        ("gap_pen",),                                    "0.00%", 12),
    ("gap_nopen_%",      ("gap_nopen",),                                  "0.00%", 13),
    ("solver_gap_%",     ("solver_gap",),                                 "0.00%", 12),
    ("oracle_mipgap_%",  ("oracle", "gap"),                               "0.00%", 14),
    ("oracle_feasible",  ("oracle", "feasible"),                          "@",     13),
    ("oracle_optimal",   ("oracle", "optimal"),                           "@",     13),
    ("oracle_status",    ("oracle", "status"),                            "@",     13),
    ("run_infeasible",   ("metrics", "run_infeasible"),                  "@",     13),
    ("n_violations",     ("metrics", "n_violations"),                    "0",     11),
    ("n_hos_viol",       ("metrics", "n_hos_violations"),                "0",     10),
    ("n_stranding",      ("metrics", "n_stranding"),                     "0",     11),
    ("n_repairs",        ("metrics", "n_repairs"),                       "0",     10),
    ("n_plan_viol",      ("metrics", "n_plan_violations"),               "0",     11),
    ("tw_n_customers",   ("metrics", "tw_n_customers"),                  "0",     13),
    ("tw_n_misses",      ("metrics", "tw_n_misses"),                     "0",     11),
    ("tw_hit_rate",      ("metrics", "tw_hit_rate"),                     "0.0%",  11),
    ("tw_penalty_h",     ("metrics", "tw_penalty_h"),                    "0.000", 12),
    ("directive_ok",     ("metrics", "directive_compliance", "compliant"), "@",  12),
    ("ro_obj",           ("ro_obj",),                                     "0.000", 11),
    ("twosp_obj",        ("twosp_obj",),                                  "0.000", 11),
    ("safety_buffer",    ("safety_buffer",),                              "0.00",  12),
    ("note",             ("note",),                                       "@",     50),
]

# run_id pattern produced by runner_dispatch.run_batch:
#   <instance>_<ALGO>_<YYYYMMDD>_<HHMMSS>_<idx>
# The batch runner appends a per-combination index ("_000"), single runs do
# not — so the index MUST be optional here.  When it was mandatory, every
# single run failed to match and fell back to a path-string ranking in
# _dedup_latest, which made stale runs beat newer ones (a path starting with
# "s"/"R" sorts above a "2026..." timestamp).
# Shared with the runner (src/paths.py), so the id the runner writes and the id
# reporting parses can never drift.  It also carries the optional VARIANT group
# — the method-configuration sweep label.
_RUN_ID_RE = _paths.RUN_ID_RE
_ALGO_TO_METHOD = {"GREEDY": "greedy", "LA": "LA", "RO": "RO",
                   "ROBU": "ROBU", "2SP": "2SP"}

_STATUS_LINE_RE = re.compile(r"Status\s*:\s*(\w+)\s*\(([\d.]+)s\)")

# Written by runner.finalize_run when paths["sol"] is None, i.e. the run is an
# internal one (the ORACLE MIP's greedy warm start) whose result must never be
# recorded as a method result.  Kept as a constant so the producer and this
# consumer cannot drift.
_INTERNAL_RUN_MARKER = "not persisted (internal warm-start run)"

# instance id -> instance family, e.g. "RshortCfewTtight_1" -> "RshortCfewTtight"
_INSTANCE_SEED_RE = re.compile(r"_\d+$")

# instance/family tag -> (route_class, customers_class, window_class):
#   R{short|medium|long}C{few|medium|many}T{none|tight|medium|large}[_<seed>]
_INSTANCE_TAG_RE = re.compile(
    r"^R(?P<route>short|medium|long)"
    r"C(?P<cust>few|medium|many)"
    r"T(?P<window>none|tight|medium|large)"
    r"(?:_\d+)?$"
)

# ── canonical orderings / display labels ────────────────────────────────────
_ROUTE_ORDER   = ["short", "medium", "long"]
_CUST_ORDER    = ["few", "medium", "many"]
_TW_ORDER      = ["none", "tight", "medium", "large"]
_ROUTE_DISPLAY = {"short": "Short", "medium": "Medium", "long": "Long"}
_CUST_DISPLAY  = {"few": "Few", "medium": "Medium", "many": "Many"}
_TW_DISPLAY    = {"none": "None", "tight": "Tight", "medium": "Medium", "large": "Large"}
# instance size class for the runtime table
_ROUTE_SIZE    = {"short": "Small", "medium": "Medium", "long": "Large"}
_SIZE_ORDER    = ["Small", "Medium", "Large"]


def _instance_family(instance: str) -> str:
    return _INSTANCE_SEED_RE.sub("", instance or "")


def _parse_instance_tags(instance: str) -> dict:
    m = _INSTANCE_TAG_RE.match(instance or "")
    if not m:
        return dict(route_class=None, customers_class=None, window_class=None)
    return dict(route_class=m.group("route"),
                customers_class=m.group("cust"),
                window_class=m.group("window"))


def _get(d: dict, path: tuple):
    """Safely walk a nested dict using a key-path tuple."""
    val = d
    for key in path:
        if not isinstance(val, dict):
            return None
        val = val.get(key)
    return val


def _safe_float(v):
    """Return float or None; treat inf/nan as None."""
    try:
        f = float(v)
        return None if (math.isnan(f) or math.isinf(f)) else f
    except (TypeError, ValueError):
        return None


def _mean(lst):
    vals = [v for v in lst if v is not None]
    return sum(vals) / len(vals) if vals else None


def _min(lst):
    vals = [v for v in lst if v is not None]
    return min(vals) if vals else None


def _max(lst):
    vals = [v for v in lst if v is not None]
    return max(vals) if vals else None


def _mode(lst):
    vals = [v for v in lst if v is not None]
    return Counter(vals).most_common(1)[0][0] if vals else None


# ── LaTeX cell formatters ────────────────────────────────────────────────────
def _fmt(x, nd=1):
    """Fixed-decimal, or '--' when missing."""
    return "--" if x is None else f"{x:.{nd}f}"


def _pct(x, nd=1):
    """Fraction -> percentage number (no % sign), or '--' when missing."""
    return "--" if x is None else f"{x * 100:.{nd}f}"


# ══════════════════════════════════════════════════════════════════════════════
# LOADING
# ══════════════════════════════════════════════════════════════════════════════

def _normalise_la_config(rows: list[dict]) -> int:
    """Re-file every row onto its EFFECTIVE variant; -> number of rows moved.

    The standard LA configuration solves its look-ahead tail as a MILP, but the
    stored corpus predates that decision: those runs carry the tag "MIPTAIL"
    and the superseded LP-tail runs carry no tag at all.  This maps both onto
    the convention the whole reporting layer already keys on — no variant means
    the standard configuration — so the base-case figures, the LaTeX tables, the
    dedup and the LA sweep cells all follow the swap without any of them knowing
    it happened.  See paths.effective_variant for the mapping.

    Applied at LOAD time, on the row dicts this module hands out, so no consumer
    can read a raw variant by accident and no file on disk is rewritten: the
    run_id remains the record of how the run was actually launched.
    """
    n = 0
    for r in rows:
        eff = _paths.effective_variant(r.get("method"), r.get("variant"),
                                       r.get("solve_mode"),
                                       r.get("la_energy_quantile"))
        if eff != (r.get("variant") or None):
            r["variant"] = eff
            n += 1
    return n


def load_solutions(solutions_dir: str) -> list[dict]:
    """Load all *.json solution files from solutions_dir (skips oracle caches).

    Goes through run_cache, so the corpus is parsed once per machine instead of
    once per reporting script, and the trajectory arrays — which nothing here
    reads — never enter memory.  See src/output_analysis/run_cache.py.
    """
    if not os.path.isdir(solutions_dir):
        print(f"  ERROR: directory not found: '{solutions_dir}'", file=sys.stderr)
        sys.exit(1)

    rows = []
    for name, data in run_cache.load_runs(solutions_dir):
        if "_error" in data:
            print(f"  SKIP {_paths.path_in(solutions_dir, name)}: "
                  f"{data['_error']}", file=sys.stderr)
            continue
        # copied: the cache hands out the record it will hand the next caller
        # in this process, and the annotators below write into these dicts
        data = dict(data)
        data["_file"] = name
        data["status"] = "OK"
        data["note"] = ""
        rows.append(data)

    if not rows:
        print(f"  WARNING: no run .json files found in '{solutions_dir}/'")
    print(f"  Loaded {len(rows)} finished run(s) from '{solutions_dir}/'")
    n_la = _normalise_la_config(rows)
    if n_la:
        print(f"  Re-filed {n_la} LA run(s) onto the standard configuration "
              f"(MILP tail) / the '{_paths.LA_LEGACY_VARIANT}' variant (LP tail)")
    return rows


def find_failed_runs(logs_dir: str, solutions_dir: str) -> list[dict]:
    """
    Find runs that have a log file but never produced a solution JSON —
    i.e. they were started but did not finish (proven infeasible, crashed,
    timed out, or interrupted).
    """
    if not os.path.isdir(logs_dir):
        print(f"  WARNING: logs directory not found: '{logs_dir}/' "
              f"(skipping unfinished-run detection)")
        return []

    # Both trees are bucketed by experiment, so these sweep the bucket
    # subdirectories as well as the root and key on the BASENAME — a run_id is
    # unique across the whole corpus, and a log is matched to its solution by
    # name, never by which bucket the two happen to sit in.
    finished_ids = {
        name[:-5] for name, _p in _paths.scan_tree(solutions_dir)
        if name.endswith(".json") and not name.startswith("oracle_")
    }

    rows = []
    n_internal = 0
    for f, log_full in _paths.scan_tree(logs_dir):
        if not f.endswith(".txt"):
            continue
        run_id = f[:-4]
        if run_id in finished_ids:
            continue

        # ORACLE runs write no solution file — their result IS the shared cache
        # solutions/oracle_<instance>.json.  Counting their logs here reported
        # every completed oracle as an "unfinished run".
        if "_ORACLE_" in run_id or run_id.endswith("_ORACLE"):
            continue

        m = _RUN_ID_RE.match(run_id)
        if m:
            instance = m.group("instance")
            method   = _ALGO_TO_METHOD.get(m.group("algo"), "UNKNOWN")
            variant  = m.group("variant")
        else:
            instance, method, variant = run_id, "UNKNOWN", None

        log_path = log_full
        try:
            text = open(log_path, encoding="utf-8", errors="replace").read()
        except Exception as e:
            text = ""
            print(f"  SKIP {log_path}: {e}", file=sys.stderr)

        # An INTERNAL run: the greedy warm start of the ORACLE MIP, which
        # completes normally but deliberately persists no solution.  Its log is
        # not evidence of a failed run — treating it as one produced a phantom
        # INCOMPLETE row that outranked (the oracle runs last) and evicted the
        # REAL greedy run of that instance from every table and figure.
        # greedy.py now writes these under logs/_internal/, which this top-level
        # scan never sees; the marker check keeps the ones already on disk out.
        if _INTERNAL_RUN_MARKER in text:
            n_internal += 1
            continue

        sm = _STATUS_LINE_RE.search(text)
        solver_status = sm.group(1) if sm else None
        elapsed       = _safe_float(sm.group(2)) if sm else None

        if solver_status == "infeasible" or "no feasible solution found" in text.lower():
            status = "INFEASIBLE"
            note = "Solver proved infeasible"
            if elapsed is not None:
                note += f" after {elapsed:.1f}s"
        else:
            status = "INCOMPLETE"
            note = "Run log ends without a completion message"
            if solver_status is not None and elapsed is not None:
                note = (f"MILP reached '{solver_status}' after {elapsed:.1f}s "
                        f"but the run crashed/was interrupted before finishing")
            note += " (crash, timeout, or manual interruption; reason not captured in the log)"

        rows.append(dict(
            run_id=run_id, instance=instance, method=method, variant=variant,
            status=status, note=note,
            wall_clock_s=elapsed,
            oracle={}, metrics={},
        ))

    # Same re-filing as the finished runs get, so an unfinished MILP-tail run
    # is not counted as an unfinished variant.  These rows carry no solve_mode
    # (there is no solution file to read it from), so the run-id tag decides.
    _normalise_la_config(rows)

    print(f"  Found {len(rows)} unfinished run(s) referenced in '{logs_dir}/' "
          f"with no matching solution file")
    if n_internal:
        print(f"  Ignored {n_internal} internal warm-start log(s) "
              f"(ORACLE's greedy seed — completed, persists no solution)")
    return rows


def _dedup_latest(rows: list[dict]) -> tuple[list[dict], int]:
    """
    When the same instance was solved several times with the same method
    (e.g. a rerun batch), keep only the LATEST run — ranked by the
    timestamp+index in the run_id (file name as fallback) — and drop the
    rest.  Keyed on (instance, method, supervised, variant) so supervised and
    unsupervised runs, and each cell of a method-configuration sweep, are kept
    separately.

    The `variant` leg is what lets a sweep run on the BASE instances: a run
    labelled e.g. "S25H12" no longer competes with — and cannot displace — the
    unlabelled base run of the same instance and method.  Every pre-existing
    run has no `variant` key, so it reads as None and this key is identical to
    the old (instance, method, supervised) one for all of them.

    NOTE: the key is deliberately NOT the stored parameter set.  1802 keys in
    the current solutions/ hold runs whose parameters differ (mostly greedy at
    prune_quantile 0.95 vs None from the guard sweep); keying on parameters
    would resurrect all of them as extra samples in the paper tables.  A run is
    a variant only when it was explicitly launched as one.

    Returns (kept_rows, n_dropped).
    """
    def _rank(r):
        """Recency key: (timestamp, batch index).  Runs whose id carries no
        parseable timestamp rank below every timestamped run (empty string
        sorts first) rather than above them by accident of path spelling."""
        m = _RUN_ID_RE.match(r.get("run_id") or "")
        if m:
            return (m.group("ts"), int(m.group("idx") or 0))
        return ("", 0)

    best: dict = {}
    n_dup = 0
    for r in rows:
        key = (r.get("instance"), r.get("method"), bool(r.get("supervised")),
               r.get("variant") or None)
        if key in best:
            n_dup += 1
            if _rank(r) > _rank(best[key]):
                best[key] = r
        else:
            best[key] = r
    return list(best.values()), n_dup


def _annotate_instance_tags(rows: list[dict]):
    """Add route_class / customers_class / window_class keys parsed from the
    instance id (e.g. "RshortCfewTtight_1" -> short / few / tight)."""
    for rec in rows:
        rec.update(_parse_instance_tags(rec.get("instance")))


# Gurobi's closing line, e.g.
#   "Best objective 1.726e+02, best bound 1.699e+02, gap 1.5735%"
# A run stopped before any incumbent prints "gap -", which parses to None.
_GUROBI_GAP_RE = re.compile(
    r"best bound\s+[-\d.e+]+,\s*gap\s+([\d.]+)%", re.IGNORECASE)


def _annotate_solver_gap(rows: list[dict], logs_dir: str):
    """Attach ``solver_gap`` — the offline MIP's own final relative gap.

    The methods store only status/optimal flags, so the distance between the
    plan they returned and the best bound their solver proved lives in the
    Gurobi log alone.  The stored ``gurobi_log`` path is the one the run wrote
    (possibly on a cluster filesystem), so only its basename is trusted and
    resolved against the local logs directory; runs whose log was not kept
    simply get None and drop out of the average.
    """
    n_hit = 0
    for rec in rows:
        rec["solver_gap"] = None
        cand = rec.get("gurobi_log") or (
            f"{rec.get('run_id')}_gurobi.log" if rec.get("run_id") else None)
        if not cand:
            continue
        path = _paths.find_in(logs_dir, os.path.basename(str(cand)))
        if path is None:
            continue
        try:
            with open(path, encoding="utf-8", errors="ignore") as fh:
                hits = _GUROBI_GAP_RE.findall(fh.read())
        except OSError:
            continue
        if hits:
            rec["solver_gap"] = float(hits[-1]) / 100.0
            n_hit += 1
    return n_hit


def _oracle_for(instance: str, solutions_dir: str):
    """The shared oracle cache for `instance`, or None when it is unsolved.

    The oracle is decoupled from method runs: a method's solution file no longer
    embeds it, so the gap is computed here from solutions/oracle_<instance>.json
    whenever it exists.  run_cache hands back the oracle's scalars plus the three
    schedule facts the gap needs (_ta_N, _misses, _n_sol) — the schedule itself
    is ~70 KB an oracle and was only ever reduced to those numbers.
    """
    rec = run_cache.load_oracles(solutions_dir).get(instance)
    return None if (rec is None or "_error" in rec) else rec


def _oracle_schedule_facts(oracle: dict) -> tuple:
    """-> (ta_N, misses) from either a cached oracle or an embedded legacy one.

    Runs written before the oracle was decoupled carry their own `oracle` block
    with a full `sol` list; those are read here so the fallback in
    _annotate_gap_to_oracle keeps working.
    """
    if "_ta_N" in oracle or "_misses" in oracle:
        return _safe_float(oracle.get("_ta_N")), oracle.get("_misses")
    sol = oracle.get("sol") or []
    if not sol:
        return None, None
    return (_safe_float(sol[-1].get("ta")),
            sum(int(s.get("delta") or 0) for s in sol))


def _annotate_gap_to_oracle(rows: list[dict], solutions_dir: str = _paths.solutions()):
    """
    Derive DURATION-based objectives and the two oracle gaps.

    The stored sim_arrival_h and oracle.obj are absolute clock values (and
    oracle.obj additionally contains the window penalty beta * misses).  Here
    both sides are converted to route durations (arrival - T_START, with
    T_START recovered as sim_arrival_h - duration_h) and compared twice:

      gap_pen   : penalised durations on both sides — (duration + beta*misses)
                  vs the oracle's (duration + beta*misses); the true
                  objective-function gap, expressed in duration terms.
      gap_nopen : pure route durations, window penalties excluded.

    Also derives oracle_duration_h / oracle_duration_pen_h / oracle_tw_misses
    (misses = sum of the delta indicators in the oracle schedule) and the
    method's penalised duration duration_pen_h.
    """
    for rec in rows:
        dur = _safe_float(rec.get("duration_h"))
        arr = _safe_float(rec.get("sim_arrival_h"))
        pen = _safe_float(_get(rec, ("metrics", "tw_penalty_h")))
        dur_pen = (dur + pen) if (dur is not None and pen is not None) else dur

        ora_dur = ora_dur_pen = ora_miss = None
        gap_pen = gap_nopen = None
        if rec.get("status") == "OK":
            t0   = (arr - dur) if (arr is not None and dur is not None) else None
            # oracle from the shared cache (decoupled); fall back to any block
            # embedded by an older run for backward compatibility
            oracle = _oracle_for(rec.get("instance"), solutions_dir) \
                     or rec.get("oracle") or {}
            oobj = _safe_float(oracle.get("obj"))
            ta_N, misses = _oracle_schedule_facts(oracle)
            if oobj is not None and t0 is not None:
                if ta_N is not None:
                    ora_miss    = misses
                    ora_dur     = ta_N - t0
                    ora_dur_pen = ora_dur + (oobj - ta_N)   # + beta * misses
            if dur is not None and ora_dur not in (None, 0):
                gap_nopen = (dur - ora_dur) / ora_dur
            if dur_pen is not None and ora_dur_pen not in (None, 0):
                gap_pen = (dur_pen - ora_dur_pen) / ora_dur_pen

        rec["duration_pen_h"]        = dur_pen
        rec["oracle_duration_h"]     = ora_dur
        rec["oracle_duration_pen_h"] = ora_dur_pen
        rec["oracle_tw_misses"]      = ora_miss
        rec["gap_pen"]               = gap_pen
        rec["gap_nopen"]             = gap_nopen
        # legacy key, consumed by the LaTeX gap table: penalty-included gap
        rec["gap_to_oracle"]         = gap_pen


# ══════════════════════════════════════════════════════════════════════════════
# RUN SELECTORS / METRIC ACCESSORS
# ══════════════════════════════════════════════════════════════════════════════

def _is_ok(r):          return r.get("status") == "OK"
def _is_supervised(r):  return bool(r.get("supervised"))

# M9 (2026-07-29) — the weekly working-time cap is out of problem scope (the
# paper models the daily provisions only), so a stored "hos_weekly" violation
# is a diagnostic, never grounds for infeasibility.  Runs written before the
# change still carry it inside metrics.violations / run_infeasible; the
# helpers below reclassify them on the fly so no re-run is needed.
_NONFATAL_VIOL = {"hos_weekly"}


def _fatal_viol_count(r) -> int:
    """Number of run-infeasibility-grounds violations (excludes diagnostics)."""
    byt = _get(r, ("metrics", "violations_by_type"))
    if isinstance(byt, dict):
        return sum(int(v) for k, v in byt.items() if k not in _NONFATAL_VIOL)
    # very old runs without the by-type dict: trust the stored flag
    return 1 if _get(r, ("metrics", "run_infeasible")) else 0


def _run_infeasible(r) -> bool:
    return (bool(_get(r, ("metrics", "run_infeasible")))
            and _fatal_viol_count(r) > 0)


def _hos_viol_count(r) -> int:
    """HoS violations excluding out-of-scope diagnostics (hos_weekly)."""
    byt = _get(r, ("metrics", "violations_by_type"))
    if isinstance(byt, dict):
        return sum(int(v) for k, v in byt.items()
                   if k.startswith("hos") and k not in _NONFATAL_VIOL)
    return int(_get(r, ("metrics", "n_hos_violations")) or 0)


def _is_feasible(r):    return not _run_infeasible(r)


# ── outcome classification ───────────────────────────────────────────────────
# Two independent axes are recorded per run so the paper never conflates a
# solver that ran out of time with a plan that genuinely fails:
#
#   solve_status  — did the OFFLINE optimiser finish?  "optimal" (proven /
#                   C&CG-converged), "time_limit" (valid incumbent returned but
#                   optimality not proven), or "n/a" for the online policies
#                   (greedy, LA), which have no offline optimum to prove.
#
#   outcome       — feasibility of the EXECUTED plan, but only counted as a
#                   genuine method failure when the plan was a certified output:
#                     "feasible"          simulated run respects every constraint
#                     "infeasible"        run fails AND the plan was certified,
#                                         so the failure is the method's own
#                                         (true infeasibility / out-of-set
#                                         realisation)
#                     "unsolved"          run fails but the offline solve never
#                                         certified the plan (ROBU C&CG did not
#                                         converge) — the failure is a solve
#                                         artifact, NOT a true infeasibility
#                     "solver_infeasible" solver proved the model infeasible
#                                         (recovered from logs/)
#                     "incomplete"        crashed / interrupted (from logs/)

def _solver_optimal(r):
    """True/False when the offline optimiser's optimality is defined for this
    method (RO, 2SP, ROBU), else None for the online policies.

    ROBU is optimal only when the C&CG loop converged AND the final master was
    proven optimal; a converged-but-time-limited master is a valid robust plan
    but not a proven optimum, so it reads as time-limited."""
    m = r.get("method")
    if m == "RO":
        return bool(r.get("ro_optimal"))
    if m == "2SP":
        return bool(r.get("twosp_optimal"))
    if m == "ROBU":
        return bool(r.get("robu_converged")) and bool(r.get("ro_optimal"))
    return None


def _plan_certified(r):
    """Is the executed plan a feasibility-complete output of the method, so that
    a simulated failure is the method's own rather than a solve artifact?

    RO / 2SP incumbents always satisfy their own model's constraints (the time
    limit costs optimality, not feasibility), so their failures are genuine.
    ROBU's plan is certified only when the cutting-plane loop converged; a
    non-converged plan still has outstanding feasibility cuts, so its stranding
    is an artifact of the aborted solve.  Online policies always execute a
    genuine plan."""
    if r.get("method") == "ROBU":
        return bool(r.get("robu_converged"))
    return True


def _solve_status(r):
    opt = _solver_optimal(r)
    if opt is None:
        return "n/a"
    return "optimal" if opt else "time_limit"


def classify_outcome(r):
    st = r.get("status")
    if st == "INFEASIBLE":
        return "solver_infeasible"
    if st != "OK":
        return "incomplete"
    # An uncertified plan (ROBU C&CG did not converge) is not a valid method
    # result at all: whether or not this particular realisation happened to
    # survive, neither its feasibility nor its (optimistic, under-cut) gap
    # should be reported as the method's — it is simply unsolved.
    if not _plan_certified(r):
        return "unsolved"
    if _run_infeasible(r):
        return "infeasible"
    return "feasible"


def _annotate_outcome(rows: list[dict]):
    """Attach solve_status and outcome to every run (used by the sheets, the
    LaTeX tables, and paper_figures)."""
    for r in rows:
        r["solve_status"] = _solve_status(r)
        r["outcome"]      = classify_outcome(r)


def _is_truly_infeasible(r):  return r.get("outcome") == "infeasible"
def _is_unsolved(r):          return r.get("outcome") == "unsolved"
def _gap_usable(r):
    """A run contributes a gap only when it produced a feasible executed plan."""
    return r.get("outcome") == "feasible"


def _method_group(r):
    """Canonical method bucket for the gap-table columns."""
    m = r.get("method")
    if m == "greedy":
        return "greedy"
    if m in ("RO", "ROBU", "2SP"):
        return m
    if m == "LA":
        crit = (r.get("criterion") or "mean").lower()
        return "LA_cvar" if crit == "cvar" else "LA_mean"
    return None


def _mval(r, key):
    return _safe_float(_get(r, ("metrics", key)))


# ══════════════════════════════════════════════════════════════════════════════
# STYLING
# ══════════════════════════════════════════════════════════════════════════════

def _style_header(cell, bg=_HEADER_BG, fg=_HEADER_FG):
    cell.font      = Font(bold=True, color=fg, name="Arial", size=10)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _BORDER


# The Results sheet is ~49 columns x ~11 500 rows, so _style_data runs on
# ~560 000 cells.  Building a fresh PatternFill/Font/Alignment per cell meant
# 1.7 M throwaway style objects; there are only a handful of distinct (bg, fg)
# pairs, so they are built once and shared.  openpyxl emits one <xf> per
# DISTINCT style either way, so the workbook is unchanged.
_ALIGN_DATA = Alignment(horizontal="center", vertical="center")
_FILL_CACHE: dict[str, PatternFill] = {}
_FONT_CACHE: dict[str, Font] = {}


def _fill(bg: str) -> PatternFill:
    f = _FILL_CACHE.get(bg)
    if f is None:
        f = _FILL_CACHE[bg] = PatternFill("solid", start_color=bg)
    return f


def _font(fg: str) -> Font:
    f = _FONT_CACHE.get(fg)
    if f is None:
        f = _FONT_CACHE[fg] = Font(name="Arial", size=10, color=fg)
    return f


def _style_data(cell, row_idx: int, num_fmt: str, status: str):
    if status in ("INFEASIBLE", "infeasible"):
        bg, fg = _INFEASIBLE_BG, _INFEASIBLE_FG
    elif status == "INCOMPLETE":
        bg, fg = _INCOMPLETE_BG, _INCOMPLETE_FG
    elif status == "unsolved":
        bg, fg = _UNSOLVED_BG, _UNSOLVED_FG
    else:
        bg, fg = (_ALT_ROW_BG if row_idx % 2 == 0 else "FFFFFF"), "000000"
    cell.fill      = _fill(bg)
    cell.font      = _font(fg)
    cell.alignment = _ALIGN_DATA
    cell.border    = _BORDER
    cell.number_format = num_fmt


def build_results_sheet(ws, rows: list[dict]):
    ws.title = "Results"
    ws.freeze_panes = "D2"

    for col_idx, (header, _, _, width) in enumerate(_COLS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        _style_header(cell)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 30

    for row_idx, rec in enumerate(rows, start=2):
        # rows that finished carry their outcome (feasible/infeasible/unsolved)
        # into the colour; only unfinished runs keep the log-derived status
        status = (rec.get("outcome", "feasible") if rec.get("status") == "OK"
                  else rec.get("status", "OK"))
        for col_idx, (_, path, num_fmt, _) in enumerate(_COLS, start=1):
            raw = _get(rec, path)
            if num_fmt == "@":
                val = "" if raw is None else str(raw)
            else:
                val = _safe_float(raw)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            _style_data(cell, row_idx, num_fmt, status)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(_COLS))}{len(rows) + 1}"


def build_summary_sheet(ws, rows: list[dict]):
    ws.title = "Summary"
    ws.freeze_panes = "F2"

    groups: dict[tuple, list] = defaultdict(list)
    for rec in rows:
        key = (_instance_family(rec.get("instance", "?")), rec.get("method", "?"))
        groups[key].append(rec)

    headers = [
        "instance_family", "route_class", "customers_class", "window_class", "method",
        "n_runs", "n_failed", "n_infeasible", "n_unsolved", "n_time_limit",
        "duration_h mean", "oracle_dur_h mean",
        "gap_pen_% mean (feas)", "gap_nopen_% mean (feas)",
        "tw_misses mean", "tw_hit_rate mean",
        "hos_viol_rate", "stranding_rate", "repair_rate",
        "decision_mean_s mean", "offline_solve_s mean", "method_time_s mean",
    ]
    col_widths = [22, 11, 13, 12, 9, 8, 9, 12, 11, 12,
                  16, 16, 19, 20, 13, 15, 13, 14, 12, 18, 18, 18]
    num_fmts   = ["@", "@", "@", "@", "@", "0", "0", "0", "0", "0",
                  "0.000", "0.000", "0.00%", "0.00%", "0.0", "0.0%",
                  "0.0%", "0.0%", "0.0%",
                  "0.0000", "0.0", "0.0"]

    for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        _style_header(cell, bg="375623", fg="FFFFFF")
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 40

    def _rate(recs, pred):
        return (sum(1 for r in recs if pred(r)) / len(recs)) if recs else None

    for ri, ((family, method), recs) in enumerate(sorted(groups.items()), start=2):
        ok_recs = [r for r in recs if _is_ok(r)]
        feas    = [r for r in ok_recs if _gap_usable(r)]
        # runs whose plan was actually certified — the denominator for the
        # robustness rates, so ROBU's non-converged strandings do not pollute it
        solved  = [r for r in ok_recs if not _is_unsolved(r)]
        n_runs     = len(recs)
        n_failed   = n_runs - len(ok_recs)
        n_infeas   = sum(1 for r in ok_recs if _is_truly_infeasible(r))
        n_unsolved = sum(1 for r in ok_recs if _is_unsolved(r))
        n_tlimit   = sum(1 for r in ok_recs if r.get("solve_status") == "time_limit")

        tags = _parse_instance_tags(family)
        data_row = [
            family, tags["route_class"], tags["customers_class"], tags["window_class"],
            method, n_runs, n_failed, n_infeas, n_unsolved, n_tlimit,
            _mean([_safe_float(_get(r, ("duration_h",))) for r in ok_recs]),
            _mean([r.get("oracle_duration_h") for r in ok_recs]),
            _mean([r.get("gap_pen") for r in feas]),
            _mean([r.get("gap_nopen") for r in feas]),
            _mean([_mval(r, "tw_n_misses") for r in ok_recs]),
            _mean([_mval(r, "tw_hit_rate") for r in ok_recs]),
            _rate(solved, lambda r: _hos_viol_count(r) > 0),
            _rate(solved, lambda r: (_mval(r, "n_stranding") or 0) > 0),
            _rate(solved, lambda r: (_mval(r, "n_repairs") or 0) > 0),
            _mean([_mval(r, "decision_time_mean_s") for r in ok_recs]),
            _mean([_mval(r, "offline_solve_time_s") for r in ok_recs]),
            _mean([_safe_float(_get(r, ("wall_clock_s",))) for r in ok_recs]),
        ]

        if n_failed == 0:
            bg = _SUMMARY_BG if ri % 2 == 0 else "FFFFFF"
        elif n_failed == n_runs:
            bg = _ALLFAIL_BG
        else:
            bg = _SOMEFAIL_BG

        for ci, (val, fmt) in enumerate(zip(data_row, num_fmts), start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill      = PatternFill("solid", start_color=bg)
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = _BORDER
            cell.number_format = fmt


# ══════════════════════════════════════════════════════════════════════════════
# LATEX TABLE BUILDERS
# ══════════════════════════════════════════════════════════════════════════════

def _present_sorted(values, order):
    s = set(values)
    return [v for v in order if v in s] + sorted(v for v in s if v not in order)


def _method_labels(rows):
    """Derive display labels for LA / SP from the actual scenario / horizon."""
    la = [r for r in rows if r.get("method") == "LA"]
    sp = [r for r in rows if r.get("method") == "2SP"]
    la_s = _mode([r.get("n_scenarios") for r in la])
    la_h = _mode([r.get("horizon_hours") for r in la])
    sp_s = _mode([r.get("n_scenarios") for r in sp])
    la_tag = (f"{int(la_s)}S-{la_h:g}h" if la_s and la_h is not None else "")
    la_lbl = f"LA ({la_tag})" if la_tag else "LA"
    cvar_lbl = f"LA-CVaR ({la_tag})" if la_tag else "LA-CVaR"
    sp_lbl = f"SP ({int(sp_s)}S)" if sp_s else "SP"
    return dict(greedy="Greedy", RO="RO", ROBU="ROBU",
                LA_mean=la_lbl, twosp=sp_lbl, LA_cvar=cvar_lbl)


def build_gap_latex(rows, metric: str = "gap_pen",
                    full_grid: bool = True,
                    secondary: str | None = None) -> list[str]:
    """
    Table 1 — gap to oracle (%) by Route × Customers × Time-Window × method.

    Gaps average over FEASIBLE (non-stranded) runs.  ``metric`` selects the
    gap definition:

      "gap_pen"   — penalised route duration on both sides (arrival +
                    beta*window-misses vs the oracle's); the objective-function
                    gap.  Oracle (h) is the mean penalised oracle duration.
      "gap_nopen" — pure route duration, window penalties excluded (beta-
                    independent efficiency gap).  Oracle (h) excludes penalties.

    ``secondary`` optionally puts a second gap definition in parentheses next
    to the main number (used for the paper table: penalised gap outside,
    duration-only gap inside).  When it is None the parentheses instead carry
    the FEASIBLE/TOTAL run count for that (class, method) — a single number
    when all runs were feasible — so sample size stays visible.
    """
    pen = (metric == "gap_pen")
    ora_field = "oracle_duration_pen_h" if pen else "oracle_duration_h"

    pop = [r for r in rows
           if _is_ok(r)
           and r.get("route_class") and r.get("customers_class")
           and r.get("window_class")]
    # CVaR column dropped; plain method headers (scenario/horizon config moves
    # to the caption instead of parenthesised headers)
    col_keys = ["greedy", "RO", "ROBU", "LA_mean", "twosp"]
    col_meth = {"greedy": "greedy", "RO": "RO", "ROBU": "ROBU",
                "LA_mean": "LA_mean", "twosp": "2SP"}
    col_hdr  = ["Greedy", "RO", "ROBU", "LA", "SP"]

    # aggregate: gaps[(route,cust,tw)][method_group] = [gaps]; oracle by instance
    gaps:    dict = defaultdict(lambda: defaultdict(list))
    gaps2:   dict = defaultdict(lambda: defaultdict(list))  # `secondary` metric
    n_total: dict = defaultdict(lambda: defaultdict(int))   # ASSESSABLE runs
    n_nobnd: dict = defaultdict(lambda: defaultdict(int))   # feasible, no oracle
    oracle:  dict = defaultdict(dict)
    n_seeds: dict = defaultdict(set)
    for r in pop:
        cell = (r["route_class"], r["customers_class"], r["window_class"])
        mg = _method_group(r)
        oobj = _safe_float(r.get(ora_field))
        if oobj is not None:
            oracle[cell][r.get("instance")] = oobj
        if mg is not None:
            # unsolved runs (plan never certified) are neither a gap sample nor
            # a genuine infeasibility — leave them out of the feas/total count
            if _is_unsolved(r):
                continue
            if _gap_usable(r) and r.get(metric) is not None:
                n_total[cell][mg] += 1
                gaps[cell][mg].append(r[metric])
                if secondary and r.get(secondary) is not None:
                    gaps2[cell][mg].append(r[secondary])
                n_seeds[cell].add(r.get("instance"))
            elif not _is_feasible(r):
                # a genuine failure: no gap exists for it BY DEFINITION, so it
                # belongs in the denominator as a missing sample
                n_total[cell][mg] += 1
            else:
                # FEASIBLE but no oracle bound cached for this instance yet:
                # the run simply cannot be assessed.  Counting it as a missing
                # gap sample (the pre-2026-07-29 behaviour) made oracle-poor
                # cells read as near-total infeasibility, e.g. "(0/50)" for
                # 50 perfectly feasible long-route runs.
                n_nobnd[cell][mg] += 1

    def _cell_str(cell, mg):
        vals = gaps[cell].get(mg, [])
        tot  = n_total[cell].get(mg, 0)
        nb   = n_nobnd[cell].get(mg, 0)
        if secondary:
            # main number = `metric`; parentheses = `secondary` on the same runs
            if not vals:
                return "--"
            sec = gaps2[cell].get(mg, [])
            body = _pct(_mean(vals), 1)
            return (rf"{body}~{{\scriptsize({_pct(_mean(sec), 1)})}}" if sec
                    else body)
        if tot == 0:
            # no assessable run; [n] flags feasible runs still awaiting an
            # oracle bound, so "no data yet" never looks like "all failed"
            return rf"--~{{\scriptsize[{nb}]}}" if nb else "--"
        nf   = len(vals)
        cnt  = f"{nf}" if nf == tot else f"{nf}/{tot}"
        body = _pct(_mean(vals), 1) if nf else "--"
        pend = rf"[{nb}]" if nb else ""
        return rf"{body}~{{\scriptsize({cnt}){pend}}}"

    # full_grid=True draws the complete canonical layout (every route,
    # customer and TW class), with '--' where no run exists yet, so the table
    # is paper-shaped from the start.
    def _custs_of(route):
        return (_CUST_ORDER if full_grid else
                _present_sorted((c[1] for c in oracle if c[0] == route),
                                _CUST_ORDER))

    def _tws_of(route, cust):
        return (_TW_ORDER if full_grid else
                _present_sorted(
                    (c[2] for c in oracle if c[0] == route and c[1] == cust),
                    _TW_ORDER))

    routes = (_ROUTE_ORDER if full_grid
              else _present_sorted((c[0] for c in oracle), _ROUTE_ORDER))
    seed_counts = sorted({len(s) for s in n_seeds.values()})
    if not seed_counts:
        n_real = r"\langle N\rangle"
    elif len(seed_counts) == 1:
        n_real = str(seed_counts[0])
    else:
        n_real = f"{seed_counts[0]}--{seed_counts[-1]}"

    # LA / SP scenario + horizon config, for the caption
    la_s = _mode([r.get("n_scenarios") for r in pop
                  if r.get("method") == "LA"])
    la_h = _mode([r.get("horizon_hours") for r in pop
                  if r.get("method") == "LA"])
    sp_s = _mode([r.get("n_scenarios") for r in pop
                  if r.get("method") == "2SP"])
    cfg_bits = []
    if la_s:
        cfg_bits.append(f"LA uses {int(la_s)} scenarios"
                        + (f" over a {la_h:g}\\,h horizon" if la_h else ""))
    if sp_s:
        cfg_bits.append(f"SP uses {int(sp_s)} scenarios")
    cfg_txt = ("  " + "; ".join(cfg_bits) + ".") if cfg_bits else ""

    pen_txt = ("window penalties included" if pen
               else "route duration only, window penalties excluded")
    # the travel-time distribution is described in the running text, not here
    if secondary:
        paren_txt = (r"  Parentheses give the same gap over the same runs "
                     r"with window penalties excluded (route duration only).  "
                     r"ROBU is solved on the short-route instances only.")
    else:
        paren_txt = (r"  Parentheses give the feasible/total run count per "
                     r"cell among runs the oracle bound can assess (a single "
                     r"number when all were feasible); square brackets give "
                     r"feasible runs still awaiting an oracle bound, which "
                     r"are excluded from both counts.")
    lbl = "tab:gap" + ("" if pen else "-nopen")
    L = []
    L.append(r"\begin{table}[htbp]")
    L.append(r"\centering")
    L.append(rf"\caption{{Gap to oracle (\%) in route duration "
             rf"({pen_txt}) by instance class and "
             rf"method, averaged over {n_real} realizations per instance "
             rf"class.{paren_txt}{cfg_txt}}}")
    L.append(rf"\label{{{lbl}}}")
    L.append(r"\resizebox{\textwidth}{!}{%")
    L.append(r"\begin{tabular}{lll c ccccc}")
    L.append(r"\toprule")
    L.append(r"\multirow{2}{*}{\textbf{\small Route}} "
             r"& \multirow{2}{*}{\textbf{\small Cust.}} "
             r"& \multirow{2}{*}{\textbf{\small TW}} "
             r"& \multirow{2}{*}{\textbf{Oracle (h)}} &")
    L.append(r"\multicolumn{5}{c}{\textbf{Gap to Oracle \%}} \\")
    L.append(r"\cmidrule(lr){5-9}")
    L.append(" & & & & " + " & ".join(col_hdr) + r" \\")
    L.append(r"\midrule")

    for ri, route in enumerate(routes):
        custs = _custs_of(route)
        route_span = sum(len(_tws_of(route, cu)) for cu in custs)
        if route_span == 0:
            continue
        first_route = True
        for cust in custs:
            tws = _tws_of(route, cust)
            if not tws:
                continue
            first_cust = True
            for tw in tws:
                cell = (route, cust, tw)
                oracle_h = _mean(list(oracle[cell].values()))
                gap_cells = [_cell_str(cell, col_meth[ck]) for ck in col_keys]
                c0 = (rf"\multirow{{{route_span}}}{{*}}{{\small {_ROUTE_DISPLAY[route]}}}"
                      if first_route else "")
                c1 = (rf"\multirow{{{len(tws)}}}{{*}}{{\small {_CUST_DISPLAY[cust]}}}"
                      if first_cust else "")
                c2 = rf"{{\small {_TW_DISPLAY.get(tw, tw)}}}"
                L.append(f"{c0} & {c1} & {c2} & {_fmt(oracle_h, 1)} & "
                         + " & ".join(gap_cells) + r" \\")
                first_route = False
                first_cust = False
        if ri != len(routes) - 1:
            L.append(r"\midrule")

    L.append(r"\bottomrule")
    L.append(r"\end{tabular}%")
    L.append(r"}")
    L.append(r"\end{table}")
    return L


def build_gap_simplified_latex(rows, metric: str = "gap_pen",
                               secondary: str | None = "gap_nopen") -> list[str]:
    """
    Simplified Table 1 — gap to oracle (%) by ROUTE x CUSTOMERS class.

    Same numbers and conventions as ``build_gap_latex``, but the time-window
    classes are pooled into the row mean instead of forming their own rows.
    Pooling is at RUN level, so each row is the mean over every assessable run
    of that (route, customers) class rather than a mean of cell means — the
    two differ whenever cells hold unequal numbers of usable runs.

    ROBU is omitted (dropped from the paper); columns run RO / 2SP / greedy /
    LA, i.e. open-loop before closed-loop.
    """
    pen = (metric == "gap_pen")
    ora_field = "oracle_duration_pen_h" if pen else "oracle_duration_h"

    pop = [r for r in rows
           if _is_ok(r)
           and r.get("route_class") and r.get("customers_class")
           and r.get("window_class")]

    col_keys = ["RO", "2SP", "greedy", "LA_mean"]
    col_hdr = ["RO", "2SP", "Greedy", "LA"]

    gaps: dict = defaultdict(lambda: defaultdict(list))
    gaps2: dict = defaultdict(lambda: defaultdict(list))
    oracle: dict = defaultdict(dict)
    for r in pop:
        cell = (r["route_class"], r["customers_class"])
        oobj = _safe_float(r.get(ora_field))
        if oobj is not None:
            oracle[cell][r.get("instance")] = oobj
        mg = _method_group(r)
        if mg is None or _is_unsolved(r):
            continue
        if _gap_usable(r) and r.get(metric) is not None:
            gaps[cell][mg].append(r[metric])
            if secondary and r.get(secondary) is not None:
                gaps2[cell][mg].append(r[secondary])

    def _cell_str(cell, mg):
        vals = gaps[cell].get(mg, [])
        if not vals:
            return "--"
        body = _pct(_mean(vals), 1)
        sec = gaps2[cell].get(mg, []) if secondary else []
        return (rf"{body}~{{\scriptsize({_pct(_mean(sec), 1)})}}" if sec
                else body)

    routes = [rt for rt in _ROUTE_ORDER
              if any(c[0] == rt for c in oracle)]

    def _custs_of(route):
        return [cu for cu in _CUST_ORDER if oracle.get((route, cu))]

    la_s = _mode([r.get("n_scenarios") for r in pop if r.get("method") == "LA"])
    la_h = _mode([r.get("horizon_hours") for r in pop if r.get("method") == "LA"])
    sp_s = _mode([r.get("n_scenarios") for r in pop if r.get("method") == "2SP"])
    cfg_bits = []
    if la_s:
        cfg_bits.append(f"LA uses {int(la_s)} scenarios"
                        + (f" over a {la_h:g}\\,h horizon" if la_h else ""))
    if sp_s:
        cfg_bits.append(f"2SP uses {int(sp_s)} scenarios")
    cfg_txt = ("  " + "; ".join(cfg_bits) + ".") if cfg_bits else ""

    n_inst = sorted({len(v) for v in oracle.values()})
    if not n_inst:
        inst_txt = "all"
    elif len(n_inst) == 1:
        inst_txt = str(n_inst[0])
    else:
        inst_txt = f"{n_inst[0]}--{n_inst[-1]}"

    pen_txt = ("window penalties included" if pen
               else "route duration only, window penalties excluded")
    paren_txt = (r"  Parentheses give the same gap over the same runs with "
                 r"window penalties excluded (route duration only)."
                 if secondary else "")

    lbl = "tab:gap-simplified" + ("" if pen else "-nopen")
    L = []
    L.append(r"\begin{table}[htbp]")
    L.append(r"\centering")
    L.append(rf"\caption{{Gap to oracle (\%) in route duration ({pen_txt}) by "
             rf"route and customer-count class and method, pooling the "
             rf"time-window classes.  Each row averages over the {inst_txt} "
             rf"instances of the class and all their realizations."
             rf"{paren_txt}{cfg_txt}}}")
    L.append(rf"\label{{{lbl}}}")
    L.append(r"\begin{tabular}{ll c cccc}")
    L.append(r"\toprule")
    L.append(r"\multirow{2}{*}{\textbf{\small Route}} "
             r"& \multirow{2}{*}{\textbf{\small Cust.}} "
             r"& \multirow{2}{*}{\textbf{Oracle (h)}} &")
    L.append(r"\multicolumn{4}{c}{\textbf{Gap to Oracle \%}} \\")
    L.append(r"\cmidrule(lr){4-7}")
    L.append(" & & & " + " & ".join(col_hdr) + r" \\")
    L.append(r"\midrule")
    for ri, route in enumerate(routes):
        custs = _custs_of(route)
        if not custs:
            continue
        first_route = True
        for cust in custs:
            cell = (route, cust)
            oracle_h = _mean(list(oracle[cell].values()))
            gap_cells = [_cell_str(cell, ck) for ck in col_keys]
            c0 = (rf"\multirow{{{len(custs)}}}{{*}}{{\small {_ROUTE_DISPLAY[route]}}}"
                  if first_route else "")
            c1 = rf"{{\small {_CUST_DISPLAY[cust]}}}"
            L.append(f"{c0} & {c1} & {_fmt(oracle_h, 1)} & "
                     + " & ".join(gap_cells) + r" \\")
            first_route = False
        if ri != len(routes) - 1:
            L.append(r"\midrule")
    L.append(r"\bottomrule")
    L.append(r"\end{tabular}")
    L.append(r"\end{table}")
    return L


def build_feasibility_latex(rows) -> list[str]:
    """
    Table 2 — feasibility / robustness by method.

    The robustness rates (infeasibility, window hit) are
    computed over CERTIFIED runs only — runs whose plan the offline solver
    actually finished — so a solver that timed out without certifying its plan
    (ROBU C&CG not converged) does not masquerade as a stranding failure.  Two
    columns describe the offline solve itself:

      Not opt. (\%)   fraction of runs where the offline optimiser returned a
                      valid plan but did not prove optimality (time limit);
                      '--' for the online policies, which have no offline
                      optimum.  These runs still count as feasible/infeasible.
      Opt. gap (\%)   mean final relative MIP gap over the NOT-optimal solves
                      only — a proven optimum contributes no gap, so averaging
                      it in would dilute the number the column is there to
                      report.  Read back from the Gurobi log; '--' where no log
                      was kept (ROBU writes none, its master solves being
                      iterative).

    Which solves count as not-optimal comes from the status the run recorded
    (ro_optimal / twosp_optimal / robu_converged), never from a route-class
    expectation: the short and medium routes do overwhelmingly close, and the
    long ones overwhelmingly do not, but where the record says otherwise the
    record wins.  Non-optimal solves whose log was not retained are simply
    unmeasured and drop out of the mean, so the caption reports how many solves
    each number rests on.

    Infeasibility pools the two execution failures — an HoS breach and a
    stranding — into one column, because that is the operational question: the
    share of runs the schedule did not survive.  It is an OR over runs, not the
    sum of the two rates, since a single run can do both and must not be counted
    twice.  The split still matters to the argument (the failure mode is a
    readout of what each architecture may change en route: the myopic policy
    breaches driving limits and never strands, the two-stage plan strands and
    almost never breaches), so it is reported per method in the caption, where
    it costs no column.

    The repair-frequency column is dropped: it applied to one method only and a
    dash in every other row spent a column on a single number.  It remains in
    the Results sheet.
    """
    ok = [r for r in rows if _is_ok(r)]

    # Plain method names: the scenario/horizon configuration belongs in the
    # caption or the running text, not in a parenthesis on every row label.
    # (data method value, display label, has offline opt?)
    method_rows = [
        ("greedy", "Greedy", False),
        ("RO",     "RO",     True),
        ("ROBU",   "ROBU",   True),
        ("LA",     "LA",     False),
        ("2SP",    "SP",     True),
    ]

    def _rate(recs, pred):
        return (sum(1 for r in recs if pred(r)) / len(recs)) if recs else None

    def _breach(r):
        return _hos_viol_count(r) > 0

    def _strand(r):
        return (_mval(r, "n_stranding") or 0) > 0

    # rows first, so the caption can state what the Opt. gap column rests on
    # and carry the per-method HoS/SOC split the merged column folds away
    body, cover, split = [], [], []
    for mval, label, has_opt in method_rows:
        u      = [r for r in ok if r.get("method") == mval]
        solved = [r for r in u if not _is_unsolved(r)]
        # OR, not sum: a run that both breaches and strands is one failed run.
        infeas = _rate(solved, lambda r: _breach(r) or _strand(r))
        hos    = _rate(solved, _breach)
        strd   = _rate(solved, _strand)
        if hos is not None and strd is not None:
            split.append(rf"{label} {_pct(hos, 1)}/{_pct(strd, 1)}")
        twh  = _mean([_mval(r, "tw_hit_rate") for r in solved])
        # the offline solves that did NOT prove optimality: the only ones with
        # a gap to optimal worth averaging
        nopt_runs = [r for r in u if r.get("solve_status") == "time_limit"]
        nopt = (len(nopt_runs) / len(u) if (has_opt and u) else None)
        meas = [r.get("solver_gap") for r in nopt_runs
                if r.get("solver_gap") is not None]
        ogap = _mean(meas) if (has_opt and meas) else None
        if has_opt and nopt_runs:
            cover.append(rf"{label} {len(meas)} of {len(nopt_runs)}")
        nopt_s = _pct(nopt, 1) if has_opt else "--"
        ogap_s = _pct(ogap, 1) if ogap is not None else "--"
        body.append(f"{label} & {_pct(infeas,1)} "
                    f"& {_pct(twh,1)} & {nopt_s} & {ogap_s} \\\\")
    cover_txt = ("  It averages the solver logs retained for "
                 + ", ".join(cover) + " such solves.") if cover else ""
    split_txt = ("  Infeasibility pools the two execution failures; as an "
                 r"HoS-breach/stranding split it is "
                 + ", ".join(split) + ".") if split else ""

    L = []
    L.append(r"\begin{table}[htbp]")
    L.append(r"\centering")
    L.append(r"\caption{Feasibility and robustness statistics by method: "
             r"the share of runs that failed in execution, through either an "
             r"HoS breach or a stranding, and the window hit rate, computed "
             r"over runs whose plan the "
             r"solver certified.  ``Not opt.'' is the share of runs returned at "
             r"the time limit without a proven optimum and ``Opt.\ gap'' the "
             r"mean final MIP gap over those solves alone, the proven optima "
             r"contributing no gap; both are undefined for the online "
             r"policies, which solve no offline program."
             + split_txt + cover_txt + r"}")
    L.append(r"\label{tab:feasibility}")
    L.append(r"\resizebox{\textwidth}{!}{%")
    L.append(r"\begin{tabular}{lcccc}")
    L.append(r"\toprule")
    L.append(r"\textbf{Method} & \textbf{Infeasible (\%)} "
             r"& \textbf{TW hit (\%)} "
             r"& \textbf{Not opt. (\%)} & \textbf{Opt. gap (\%)}\\")
    L.append(r"\midrule")
    L += body
    L.append(r"\bottomrule")
    L.append(r"\end{tabular}%")
    L.append(r"}")
    L.append(r"\end{table}")
    return L


def build_runtime_latex(rows) -> list[str]:
    """
    Table 3 — offline solve (s) and per-stop decision (s) by method and
    instance size class (Small/Medium/Large = short/medium/long routes).
    """
    ok = [r for r in rows if _is_ok(r)]
    labels = _method_labels(ok)

    def _by_size(recs, key):
        out = {}
        for size in _SIZE_ORDER:
            sub = [r for r in recs
                   if _ROUTE_SIZE.get(r.get("route_class")) == size]
            out[size] = _mean([_mval(r, key) for r in sub])
        return out

    # (data method value, display, offline?, decision?)
    #
    # has_dec=False is a MODELLING statement, not a missing measurement: RO and
    # ROBU commit every activity and every duration before departure and the
    # simulator replays the plan, so there is no per-stop decision to time.
    # Printing 0.00 there invited the reading that they decide at every stop and
    # happen to be fast, which is the opposite of the architectural point the
    # section makes.  Greedy does decide at every stop and is merely
    # instantaneous, so it keeps its zeros.
    #
    # The oracle is not a policy but an ex-post bound (Section 6.7), so it has
    # no online cost by construction and its offline solve time is not retained
    # in the cache; a row of six dashes carried no information and is dropped.
    method_rows = [
        ("greedy", labels["greedy"], False, True),
        ("RO",     labels["RO"],     True,  False),
        ("ROBU",   labels.get("ROBU", "ROBU"), True,  False),
        ("LA",     labels["LA_mean"], False, True),
        ("2SP",    labels["twosp"],  True,  True),
    ]

    L = []
    L.append(r"\begin{table}[htbp]")
    L.append(r"\centering")
    L.append(r"\caption{Computation times by method and instance class: "
             r"offline solve time (s) and per-stop online decision time (s).  "
             r"Every offline solve is capped at a 7\,200\,s time limit, so a "
             r"figure at that value is a method that did not terminate rather "
             r"than one that took exactly that long.  ``--'' in the per-stop "
             r"columns marks a method that commits its full schedule before "
             r"departure and therefore takes no decision en route.}")
    L.append(r"\label{tab:runtime}")
    L.append(r"\begin{tabular}{lcccccc}")
    L.append(r"\toprule")
    L.append(r"& \multicolumn{3}{c}{\textbf{Offline solve (s)}} "
             r"& \multicolumn{3}{c}{\textbf{Per-stop decision (s)}}\\")
    L.append(r"\cmidrule(lr){2-4}\cmidrule(lr){5-7}")
    L.append(r"\textbf{Method} & Small & Medium & Large & Small & Medium & Large\\")
    L.append(r"\midrule")

    for mval, label, has_off, has_dec in method_rows:
        recs = [r for r in ok if r.get("method") == mval]
        off = _by_size(recs, "offline_solve_time_s") if has_off else None
        dec = _by_size(recs, "decision_time_mean_s") if has_dec else None
        off_c = ([_fmt(off[s], 1) for s in _SIZE_ORDER] if has_off
                 else ["--", "--", "--"])
        dec_c = ([_fmt(dec[s], 2) for s in _SIZE_ORDER] if has_dec
                 else ["--", "--", "--"])
        L.append(f"{label} & " + " & ".join(off_c) + " & "
                 + " & ".join(dec_c) + r" \\")

    L.append(r"\bottomrule")
    L.append(r"\end{tabular}")
    L.append(r"\end{table}")
    return L


def write_latex_sheet(ws, title: str, lines: list[str]):
    """Dump LaTeX source, one line per cell down column A (copy-paste ready)."""
    ws.title = title
    ws.column_dimensions["A"].width = 120
    mono = Font(name="Consolas", size=10)
    for i, line in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=line)
        cell.font = mono
        cell.alignment = Alignment(horizontal="left", vertical="center")


# ══════════════════════════════════════════════════════════════════════════════
# DRIVER
# ══════════════════════════════════════════════════════════════════════════════

def compile_to_excel(solutions_dir: str, logs_dir: str, output_path: str,
                     tex_dir: str | None = None):
    rows = load_solutions(solutions_dir)
    rows += find_failed_runs(logs_dir, solutions_dir)
    _annotate_instance_tags(rows)
    _annotate_gap_to_oracle(rows, solutions_dir)
    _annotate_outcome(rows)
    n_gl = _annotate_solver_gap(rows, logs_dir)
    print(f"  Read a final MIP gap from {n_gl} Gurobi log(s)")

    rows, n_dup = _dedup_latest(rows)
    if n_dup:
        print(f"  Dropped {n_dup} superseded duplicate run(s) "
              f"(same instance + method, older timestamp)")

    rows.sort(key=lambda r: (str(r.get("instance")), str(r.get("method")),
                             str(r.get("variant") or ""), str(r.get("run_id"))))

    # Base-case rows only for every AGGREGATE output.  A method-configuration
    # sweep (--variant) runs on the base instances, so unlike the "__tag"
    # instance variants it is NOT filtered out by the instance-name regex: a
    # variant row carries a perfectly valid route/customers/window class and
    # would silently pool into the published means.  The full row set still
    # goes to the Results sheet, which is the complete record of what was run.
    base_rows = [r for r in rows if not r.get("variant")]
    n_var     = len(rows) - len(base_rows)
    if n_var:
        _seen = sorted({str(r.get("variant")) for r in rows if r.get("variant")})
        print(f"  Excluded {n_var} method-variant run(s) from the summary and "
              f"LaTeX tables (kept in 'Results'): {', '.join(_seen)}")

    wb = openpyxl.Workbook()
    build_results_sheet(wb.active, rows)
    build_summary_sheet(wb.create_sheet(), base_rows)

    # LaTeX tables — penalised (objective) gap and duration-only gap
    gap_pen_tab = build_gap_latex(base_rows, metric="gap_pen",
                                  secondary="gap_nopen")
    write_latex_sheet(wb.create_sheet(), "LaTeX_Gap", gap_pen_tab)
    gap_np_tab  = build_gap_latex(base_rows, metric="gap_nopen")
    write_latex_sheet(wb.create_sheet(), "LaTeX_Gap_nopen", gap_np_tab)
    gap_simp_tab = build_gap_simplified_latex(base_rows, metric="gap_pen",
                                              secondary="gap_nopen")
    write_latex_sheet(wb.create_sheet(), "LaTeX_Gap_simple", gap_simp_tab)

    feas = build_feasibility_latex(base_rows)
    write_latex_sheet(wb.create_sheet(), "LaTeX_Feasibility", feas)

    runtime = build_runtime_latex(base_rows)
    write_latex_sheet(wb.create_sheet(), "LaTeX_Runtime", runtime)

    wb.save(output_path)

    if tex_dir:
        os.makedirs(tex_dir, exist_ok=True)
        _dump = {"gap.tex": gap_pen_tab,
                 "gap_nopen.tex": gap_np_tab,
                 "gap_simplified.tex": gap_simp_tab,
                 "feasibility.tex": feas,
                 "runtime.tex": runtime}
        for name, lines in _dump.items():
            with open(os.path.join(tex_dir, name), "w", encoding="utf-8") as fh:
                fh.write("\n".join(lines) + "\n")
        print(f"  LaTeX .tex  : {len(_dump)} file(s) -> '{tex_dir}/'")

    n_failed = sum(1 for r in rows if r.get("status") != "OK")
    sheets = ["Results", "Summary", "LaTeX_Gap", "LaTeX_Gap_nopen",
              "LaTeX_Gap_simple", "LaTeX_Feasibility", "LaTeX_Runtime"]
    print(f"  Excel saved : {output_path}")
    print(f"  Sheets      : {', '.join(sheets)}  "
          f"({len(rows)} rows, {n_failed} unfinished)")


# ── CLI ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Compile solution JSON files (plus unfinished runs found "
                     "in logs/) into an Excel summary with LaTeX-ready tables."
    )
    parser.add_argument(
        "--dir", default=_paths.solutions(),
        help="Directory containing solution .json files (default: 'solutions')"
    )
    parser.add_argument(
        "--logs", default=_paths.logs(),
        help="Directory containing run log .txt files, used to detect runs "
             "that never produced a solution file (default: 'logs')"
    )
    parser.add_argument(
        "--out", default=_paths.data_output("solution_summary.xlsx"),
        help="Output Excel file path (default: 'solution_summary.xlsx')"
    )
    # Defaults to tex/tables/ rather than None: with an opt-in flag a plain
    # run refreshed the Excel but silently left the .tex tables stale, which
    # is how the published gap table drifted from solutions/.
    parser.add_argument(
        "--tex-dir", default=_paths.tex_tables(),
        help="Directory for the generated LaTeX tables "
             f"(default: {_paths.tex_tables()}); pass '' to skip"
    )
    args = parser.parse_args()
    compile_to_excel(args.dir, args.logs, args.out, tex_dir=args.tex_dir)
